"""
Chateau in the Woods -- Five-Year Cash Flow Model builder.

Reads the three case-study source files and writes a fully formula-driven
Excel model. Every projection cell is a formula referencing the Assumptions
sheet, so changing a driver recalculates the whole model.

Sources:
  1. rent_roll20260701_1.xlsx                              (as of 07/01/2026)
  2. Chateau_Trailing_Profit_And_Loss_Detail_01312026.xlsx  (T-12 Feb'25-Jan'26)
  3. Chateau_in_the_Woods_Offering_Memorandum.pdf           (CBRE, 9/10/2025)
"""

import collections
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

UP = "/root/.claude/uploads/6b6b3d13-f8fe-539d-bb06-0a47e44c9f31"
RR_SRC = f"{UP}/0bb20064-rent_roll20260701_1.xlsx"
T12_SRC = f"{UP}/441e234e-Chateau_Trailing_Profit_And_Loss_Detail_01312026.xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Chateau_in_the_Woods_5yr_Cash_Flow_Model.xlsx")

# ----------------------------------------------------------------------------
# Styling
# ----------------------------------------------------------------------------
FONT = "Arial"
BLUE = Font(name=FONT, size=10, color="0000FF")          # hardcoded input / lever
BLACK = Font(name=FONT, size=10)                          # formula
GREEN = Font(name=FONT, size=10, color="008000")          # link to another sheet
BOLD = Font(name=FONT, size=10, bold=True)
BOLD_GREEN = Font(name=FONT, size=10, bold=True, color="008000")
ITAL = Font(name=FONT, size=9, italic=True, color="595959")
H1 = Font(name=FONT, size=14, bold=True, color="1F3864")
H2 = Font(name=FONT, size=11, bold=True, color="FFFFFF")
NOTE = Font(name=FONT, size=9, color="595959")

FILL_H2 = PatternFill("solid", fgColor="1F3864")
FILL_SUB = PatternFill("solid", fgColor="D9E1F2")
FILL_KEY = PatternFill("solid", fgColor="FFFF00")
FILL_TOT = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin", color="BFBFBF")
TOP = Border(top=Side(style="thin", color="404040"))
TOPBOT = Border(top=Side(style="thin", color="404040"),
                bottom=Side(style="double", color="404040"))
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR = '$#,##0;($#,##0);"-"'
CUR2 = '$#,##0.00;($#,##0.00);"-"'
NUM = '#,##0;(#,##0);"-"'
NUM1 = '#,##0.0;(#,##0.0);"-"'
PCT = '0.0%;(0.0%);"-"'
PCT2 = '0.00%;(0.00%);"-"'
MULT = '0.00"x"'
DATE = "mmm-yy"

YR_COLS = ["C", "D", "E", "F", "G"]      # Y1..Y5 everywhere
EXIT_COL = "H"                            # Y6 forward NOI


def w(ws, cell, value, font=BLACK, fmt=None, fill=None, align=None, border=None):
    c = ws[cell]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if align:
        if align == "wrap":
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        else:
            c.alignment = Alignment(horizontal=align, vertical="center")
    if border:
        c.border = border
    return c


def section(ws, row, title, last_col="H"):
    ws.merge_cells(f"A{row}:{last_col}{row}")
    w(ws, f"A{row}", title, H2, fill=FILL_H2, align="left")
    for col in range(1, openpyxl.utils.column_index_from_string(last_col) + 1):
        ws.cell(row=row, column=col).fill = FILL_H2
    return row + 1


def title_block(ws, subtitle, last_col="H"):
    w(ws, "A1", "CHATEAU IN THE WOODS  |  4020 Monaco Drive, Indianapolis, IN 46220  |  118 Units", H1)
    w(ws, "A2", subtitle, Font(name=FONT, size=10, bold=True, color="404040"))
    w(ws, "A3", "Blue = hardcoded input / lever   ·   Black = formula   ·   Green = link to another sheet   "
                "·   Yellow fill = key assumption", ITAL)


# ----------------------------------------------------------------------------
# Read source data
# ----------------------------------------------------------------------------
def read_rent_roll():
    ws = openpyxl.load_workbook(RR_SRC, data_only=True)["Sheet1"]
    rows = []
    for r in ws.iter_rows(min_row=11, max_row=128, values_only=True):
        if r[0] is None:
            continue
        rows.append(dict(
            unit=r[0], bdba=r[1], utype=r[2], sf=int(r[3]), status=r[4],
            mkt=float(r[5]), rent=float(r[6]) if r[6] not in (None, "") else 0.0,
            lf=r[8], lt=r[9]))
    return rows


def read_t12():
    """Return {account_label: t12_total} for the T-12 P&L detail."""
    ws = openpyxl.load_workbook(T12_SRC, data_only=True)["Sheet"]
    out = {}
    for r in ws.iter_rows(values_only=True):
        label = " ".join(str(c) for c in r[0:5] if c)
        if label and isinstance(r[17], (int, float)):
            out[label] = float(r[17])
    return out


RR = read_rent_roll()
T12 = read_t12()

WD_TYPES = {t for t in {x["utype"] for x in RR} if "WD" in t}

# floor-plan roll-up, ordered by bed/bath then rent
mix = collections.OrderedDict()
for x in RR:
    k = (x["bdba"], x["utype"], x["sf"], x["mkt"])
    d = mix.setdefault(k, dict(n=0, occ=0, rent=0.0))
    d["n"] += 1
    d["rent"] += x["rent"]
    if "vacant" not in x["status"].lower():
        d["occ"] += 1
MIX = sorted(mix.items(), key=lambda kv: (kv[0][0], kv[0][3]))

UNITS = len(RR)
NRSF = sum(x["sf"] for x in RR)
MKT_TOT = sum(x["mkt"] for x in RR)
INPLACE_TOT = sum(x["rent"] for x in RR)
OCC_UNITS = sum(1 for x in RR if "vacant" not in x["status"].lower())
N_WD = sum(1 for x in RR if x["utype"] in WD_TYPES)
N_NOWD = UNITS - N_WD

wb = openpyxl.Workbook()

# ============================================================================
# 1. README
# ============================================================================
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 108
title_block(ws, "Five-Year Cash Flow Model — Read Me, Deliverable Index & Material Assumptions", "B")

r = 5
r = section(ws, r, "HOW THIS MODEL WORKS", "B")
for a, b in [
    ("Single source of drivers",
     "Every economic assumption lives on the 'Assumptions' tab. All projection cells are formulas; changing a "
     "driver recalculates the model end-to-end, including the returns and both sensitivity grids."),
    ("Model periodicity",
     "Annual (Year 1 – Year 5) operating model, driven by a 60-month renovation / downtime schedule on the "
     "'Reno Schedule' tab. Month 1 = August 2026 (rent roll is dated 07/01/2026)."),
    ("Revenue engine",
     "Gross potential rent is built unit-by-unit from the 07/01/2026 rent roll: classic units carry in-place "
     "market rent, renovated units carry market rent plus the renovation premium, and units convert as the "
     "renovation schedule completes them."),
    ("Data hierarchy",
     "Per the case instructions, the Excel rent roll and T-12 govern wherever they conflict with the Offering "
     "Memorandum. The OM is used only for facts the Excel files do not contain (unit W/D inventory, rent comps, "
     "tax analysis, physical condition, other-income policy detail)."),
    ("Capital funding convention",
     "All renovation and property capital is funded at closing through Sources & Uses (65% loan / 35% equity). "
     "Annual cash flow therefore deducts interest and replacement reserves only, and does not double-count "
     "capex. This is conservative — it puts the full equity out on day one with no timing benefit."),
]:
    w(ws, f"A{r}", a, BOLD)
    w(ws, f"B{r}", b, BLACK, align="wrap")
    ws.row_dimensions[r].height = 30
    r += 1

r += 1
r = section(ws, r, "TAB INDEX", "B")
for a, b in [
    ("Assumptions", "All inputs and levers. Purchase price, growth, renovation scope, expenses, debt, exit."),
    ("Rent Roll", "All 118 units from the 07/01/2026 rent roll, with case status rules and per-unit renovation scope applied."),
    ("Unit Mix", "Floor-plan roll-up: in-place market rents, W/D inventory, renovation cost and premium, renovated rents."),
    ("T-12 Recast", "T-12 (Feb-25 to Jan-26) actuals mapped to underwriting line items, with the Year-1 underwritten figure and variance."),
    ("Other Income", "Line-by-line other-income build off T-12 actuals plus the identified ancillary upside, with implementation ramps."),
    ("Capex Budget", "Non-unit capital: deferred maintenance, building systems and amenity scope, phased by year."),
    ("Reno Schedule", "60-month renovation schedule: units completed, capex spend, and renovation downtime vacancy."),
    ("Annual Model", "The five-year cash flow: revenue build, operating expenses, NOI, capital and cash flow to equity."),
    ("Returns", "Sources & Uses, debt, exit valuation, unlevered and levered IRR / equity multiple / cash-on-cash."),
    ("Sensitivities", "Live grids: purchase price × exit cap (IRR and multiple) and renovation premium × exit cap."),
    ("DD Questions", "Diligence request list, with the placeholder answer assumed in this model for each open item."),
]:
    w(ws, f"A{r}", a, BOLD)
    w(ws, f"B{r}", b, BLACK, align="wrap")
    r += 1

r += 1
r = section(ws, r, "MATERIAL ASSUMPTIONS — NOT PROVIDED IN THE CASE MATERIALS", "B")
w(ws, f"A{r}", "Item", BOLD, fill=FILL_SUB)
w(ws, f"B{r}", "Assumption and basis", BOLD, fill=FILL_SUB)
r += 1
for a, b in [
    ("Purchase price — the answer, not an input",
     "$12,000,000 ($101,695/unit). The case states no price, so the model SOLVES for one: the highest basis at "
     "which the five-year plan clears a 14% levered IRR and a 1.80x multiple on the conservative case. For "
     "reference the asset actually traded on 12 June 2026 for $13,625,000 ($115,466/unit) per the Marion "
     "County deed record (special warranty deed, assessor validity flag 'Y'), 12% above our bid, where the "
     "same plan returns 7.9% and 1.42x."),
    ("Renovation scope, premium and the two-track programme",
     "TRACK A — the 53 units that already have laundry are renovated IN PLACE at 10 a month from month 2: no "
     "permit (cosmetic work is exempt in Marion County), no vacancy loss, $10,000/unit for a lighter scope, a "
     "$100/month premium captured at the next renewal (6-month lag). TRACK B — the 65 units gaining laundry "
     "need a state Construction Design Release plus a local permit AND a vacant unit, so they run at natural "
     "turnover, 5 a month from month 4, $13,000 interior plus the case-prescribed $5,000 laundry addition, a "
     "$200/month premium captured immediately, 14 days of incremental downtime. Blended $14,407/unit at "
     "$155/month = 12.9% return on cost, all 118 units complete in month 16 without forcing a single resident "
     "out. A single-track plan at 5 a month would take 27 months for the same result."),
    ("Renovated rent support",
     "Renovated rents of $1,250 (1x1) to $1,725 (3x2) sit inside the OM rent comp set. Avalon Lake (1974, same "
     "vintage, Graham Rd & Binford) achieves $1,178 / $1,655 / $1,834 for 1/2/3-bedrooms; submarket averages are "
     "$1,044 / $1,305 / $1,597. CBRE's own light-touch Upgrade Forecast is $1,094-$1,631 with a $50 premium and "
     "$1,375 on the Verdun with a W/D added."),
    ("Permit path",
     "Indiana permits multifamily in two sequential steps: multifamily is a Class 1 structure requiring a "
     "STATE Construction Design Release from IDHS (10-business-day clock, 15-20 business days typical first "
     "review) BEFORE the local Marion County DBNS permit (5-20 days) can issue — roughly 8-12 weeks from "
     "closing once design time is included, which is why Track B starts in month 4. Cosmetic work is exempt, "
     "which is what lets Track A start in month 2. On a five-year hold permit timing is not a returns risk; "
     "the risk is scope discovery — if the 65 units need panel or riser upgrades the $5,000 allowance breaks."),
    ("Market rent growth",
     "1.5% / 2.0% / 2.5% / 3.0% / 3.0%, applied on a mid-year convention. Set off ACTUALS rather than a "
     "forecast: Yardi Matrix 1Q-2026 reports Indianapolis rents FLAT on a trailing-three-month basis at $1,310 "
     "and up only 1.1% year over year, and 2025 construction starts MORE THAN DOUBLED versus 2024 — that wave "
     "delivers in 2027-28, inside our hold, with Carmel alone taking more than half of 2026 metro deliveries "
     "immediately north of this asset. Year 1 sits slightly above the 1.1% actual, Year 2 is held down by the "
     "delivery wave, and 3.0% is reached only once it clears. A flat 3.0% would add roughly 3 points of "
     "levered IRR and is not supportable on current evidence."),
    ("Real estate taxes — corrected against the public record",
     "$137,000 in Year 1 growing 4% to $160,300 by Year 5, set off the ACTUAL Marion County bill. Parcel "
     "8000002 paid $129,648 in 2025 on a $5,214,932 net assessed value at 2.4861%. The 25-year record on the "
     "'Tax Record' tab shows assessed value range-bound at $5.1-6.6mm every year since 2012 (0.36% a year), "
     "the last assessment change recorded 15 March 2022, and — decisively — assessed value FELL the year after "
     "the March 2006 sale. Indiana values apartments at the lowest of the income, market and cost approaches, "
     "which is why assessed value does not track trade prices here. An earlier draft stepped taxes to a share "
     "of the purchase price; the record does not support that, and correcting it is worth about 1.5 points of "
     "levered IRR. The reassessment risk is still priced explicitly in Ladder C on the Sensitivities tab."),
    ("Stabilised vacancy",
     "5.5% operating physical vacancy (94.5% occupancy) plus renovation downtime, against a 94.0% metro figure "
     "and a 6.8% T-12 actual. CBRE forecasts 5.0%. The half-point of conservatism reflects the 2027-28 delivery "
     "wave and Carmel's share of it."),
    ("Insurance",
     "$95,000 in Year 1 ($805/unit) growing 5%, versus $94,611 T-12 actual and CBRE's $88,326 forecast. No "
     "current loss run or bindable quote was provided."),
    ("Bad debt",
     "1.50% of GPR in Year 1 falling to 1.00% by Year 4. The T-12 shows $28,229 of gross bad debt against "
     "$19,050 of collections (net $9,179, 0.5%), but write-offs are accelerating (Jan-26 alone was $5,670) and "
     "the rent roll carries 4 units in eviction and 8 on notice."),
    ("Exit",
     "6.75% cap rate on forward (Year 6) NOI, 1.5% cost of sale, at the end of Year 5. Fifteen basis points "
     "inside the 6.90% going-in cap, reflecting a fully renovated asset with in-unit laundry throughout versus "
     "the classic condition being acquired."),
    ("Debt",
     "65% loan-to-cost on price plus all capital, 5.85% fixed, interest-only, 5-year term, 1% origination. "
     "Sized to a 10.4% Year-1 debt yield and a 1.38x Year-1 DSCR. No financing terms were provided."),
    ("A number we could not source",
     "An earlier draft compared our renovated rents to a '$1.40 per square foot' metro average. Yardi Matrix "
     "publishes a $1,310 average asking rent per unit but not a per-foot figure, so that comparison was an "
     "inference presented as data and has been removed. What can be said from the source: our Year-5 portfolio "
     "average of $1,627 per unit is about 10% above the $1,310 metro average grown at the same rates, which is "
     "what you would expect for a property whose units average 1,139 SF against a materially smaller metro "
     "average."),
    ("Excluded upside",
     "Not underwritten: converting the Building 4042 maintenance storage into two additional units, and "
     "converting 9 carports to garages. Both are real but return roughly 5-10% on cost, below the interior "
     "programme, and neither is needed for the recommendation."),
]:
    w(ws, f"A{r}", a, BOLD, fill=FILL_KEY)
    w(ws, f"B{r}", b, BLACK, align="wrap")
    ws.row_dimensions[r].height = 46
    r += 1

r += 1
r = section(ws, r, "UNDERWRITING POSTURE", "B")
for a, b in [
    ("Conservative on what we do not control",
     "Rent growth, exit cap, real estate taxes and insurance are all set at or below observable market "
     "evidence. These are forecasts, not decisions, and paying for them is how a basis gets away from you."),
    ("Confident on what we do control",
     "The renovation premium, the ancillary income build, expense management and the renovation pace are "
     "underwritten at what a competent operator should deliver — not haircut for the sake of it. The premium "
     "is still sensitised from $125 to $190 because it is unproven on this asset."),
    ("The model answers a bid, not a justification",
     "It solves for the highest price at which the business plan clears a 14% levered IRR and a 1.8x multiple "
     "under the conservative case. It is not reverse-engineered to make the asking price work. On this deal "
     "those two questions give very different answers, which is the whole finding."),
    ("Expect to lose the deal",
     "At $11.25mm this is a 21% discount to the implied ask. We should expect to be outbid by a buyer with a "
     "lower cost of capital or a more aggressive rent view. Losing at the right price is a better outcome than "
     "winning at the wrong one — the renovation cannot fix a basis mistake."),
]:
    w(ws, f"A{r}", a, BOLD)
    w(ws, f"B{r}", b, BLACK, align="wrap")
    ws.row_dimensions[r].height = 40
    r += 1

r += 1
r = section(ws, r, "CONFLICTS BETWEEN SOURCES — AND HOW THEY WERE RESOLVED", "B")
for a, b in [
    ("In-unit washer/dryer count",
     "The OM says '78 out of 118 units have a full size or stackable washer/dryer' but then lists only 53 "
     "side-by-side plus 1 stackable = 54, and its own floor-plan and building-mix schedules foot to 54 units "
     "with 'WD' in the unit code. The rent roll shows 53 units in a 'w/ WD' unit type (Rhone 15, Roman 4, "
     "Cabernet 4, Brunswick 30). Per the data hierarchy the rent roll governs: 53 units have laundry, 65 do not "
     "and carry the $5,000 addition. This is a confirmatory diligence item."),
    ("Bordeaux washer/dryer unit",
     "The OM identifies one Bordeaux (4017-D) with a stackable washer/dryer at a $1,307 market rent. The rent "
     "roll shows all 25 Bordeaux units at a $1,260 market rent in a single 'Bordeaux' unit type. The rent roll "
     "governs, so 4017-D is underwritten as needing the $5,000 addition — conservative by $5,000."),
    ("Market rents",
     "OM floor-plan mix (dated 09/08/2025) shows $1,044 / $1,109 / $1,174 / $1,343 / $1,212 / $1,257 / $1,581. "
     "The 07/01/2026 rent roll shows $1,050 / $1,125 / $1,165 / $1,350 / $1,250 / $1,260 / $1,600. The rent roll "
     "governs; it is also nine months newer."),
    ("Occupancy",
     "The OM shows 105 occupied / 13 vacant as of 09/08/2025. The 07/01/2026 rent roll shows 114 occupied / 4 "
     "vacant (96.6%) after applying the case rules that 'Notice' and 'Evict' count as occupied. The rent roll "
     "governs."),
    ("Trailing twelve months",
     "The OM's T-12 is August 2024 - July 2025 ($1,774,433 revenue / $799,669 NOI). The Excel T-12 provided is "
     "February 2025 - January 2026 ($1,790,101 revenue / $772,374 NOI on the property's own basis). The Excel "
     "T-12 governs and is recast on the 'T-12 Recast' tab."),
]:
    w(ws, f"A{r}", a, BOLD)
    w(ws, f"B{r}", b, BLACK, align="wrap")
    ws.row_dimensions[r].height = 46
    r += 1

# ============================================================================
# 2. ASSUMPTIONS
# ============================================================================
wsa = wb.create_sheet("Assumptions")
wsa.sheet_view.showGridLines = False
wsa.column_dimensions["A"].width = 46
wsa.column_dimensions["B"].width = 14
for c in YR_COLS:
    wsa.column_dimensions[c].width = 12
wsa.column_dimensions["H"].width = 62
title_block(wsa, "Underwriting Assumptions & Drivers")

A = {}          # key -> absolute ref string like "Assumptions!$B$12"
AROW = {}       # key -> row number (for year-indexed rows)


def a_val(row, label, value, fmt, note="", font=BLUE, key=None, fill=None):
    w(wsa, f"A{row}", label, BLACK)
    w(wsa, f"B{row}", value, font, fmt, fill=fill, align="center")
    if note:
        w(wsa, f"H{row}", note, NOTE, align="wrap")
    if key:
        A[key] = f"Assumptions!$B${row}"
        AROW[key] = row
    return row + 1


def a_years(row, label, values, fmt, note="", key=None, fill=None):
    w(wsa, f"A{row}", label, BLACK)
    for i, col in enumerate(YR_COLS):
        v = values[i]
        w(wsa, f"{col}{row}", v, BLUE if not isinstance(v, str) else BLACK, fmt,
          fill=fill, align="center")
    if note:
        w(wsa, f"H{row}", note, NOTE, align="wrap")
    if key:
        AROW[key] = row
    return row + 1


def yr_hdr(row):
    w(wsa, f"A{row}", "", BOLD)
    for i, col in enumerate(YR_COLS):
        w(wsa, f"{col}{row}", f"Year {i+1}", BOLD, fill=FILL_SUB, align="center")
    return row + 1


r = 5
r = section(wsa, r, "TRANSACTION")
r = a_val(r, "Purchase Price (our bid)", 12000000, CUR,
          "THIS IS THE ANSWER THE MODEL SOLVES FOR, not an input we were given. It is the maximum price at "
          "which the five-year plan clears a 14% levered IRR and a 1.80x multiple on the conservative case. "
          "See the Sensitivities tab for the live price ladder and breakeven table.", key="price", fill=FILL_KEY)
r = a_val(r, "Memo: Actual Trade Price", 13625000, CUR,
          "Marion County deed dated 06/12/2026, recorded 06/26/2026, special warranty deed, assessor validity "
          "flag 'Y' (arm's length). $115,466/unit. Prior trade 03/28/2006 at $4,400,000 was flagged 'N' "
          "(not a valid market sale). Shown for benchmarking only — it does not drive the model.",
          key="actual", fill=FILL_KEY)
r = a_val(r, "  Our Bid vs Actual Trade", None, PCT, font=BLACK, key="vs_trade")
r = a_val(r, "  Price per Unit", f"={A['price']}/{{units}}", CUR, font=BLACK, key="ppu")
r = a_val(r, "  Price per Net Rentable SF", f"={A['price']}/{{nrsf}}", CUR2, font=BLACK, key="ppsf")
r = a_val(r, "Acquisition Costs (% of price)", 0.015, PCT,
          "Title, survey, legal, third-party reports, transfer costs.", key="acqcost")
r = a_val(r, "Analysis Start (Month 1)", "2026-08-01", None,
          "Rent roll is dated 07/01/2026; assumes an August 2026 close.", key="start")
wsa[f"B{AROW['start']}"].number_format = "mmm-yyyy"
r = a_val(r, "Hold Period (years)", 5, NUM, key="hold")
r = a_val(r, "Exit Cap Rate (on Year-6 NOI)", 0.0675, PCT2,
          "15 bps inside the going-in cap: a fully renovated asset with in-unit laundry throughout trades "
          "tighter than the classic product being acquired.", key="exitcap", fill=FILL_KEY)
r = a_val(r, "Cost of Sale (% of gross value)", 0.015, PCT, key="cos")
r = a_val(r, "  Implied Going-In Cap (our Year-1 NOI)", None, PCT2, font=GREEN, key="goingin",
          note="On our underwriting, not the broker's forecast. CBRE markets a 6.9% Year-1 cap rate on its own "
               "$980,838 Acquisition Forecast NOI, which is what implies the $14.2mm price above.")

r += 1
r = section(wsa, r, "PROPERTY — FROM RENT ROLL 07/01/2026")
r = a_val(r, "Total Units", UNITS, NUM, "Rent roll total row.", key="units")
r = a_val(r, "Net Rentable Square Feet", NRSF, NUM, "Rent roll total row; ties to the OM.", key="nrsf")
r = a_val(r, "  Average Unit Size (SF)", f"={A['nrsf']}/{A['units']}", NUM, font=BLACK)
r = a_val(r, "Year Built", 1974, "0", "OM Property Summary.", key="ybuilt")
r = a_val(r, "Occupied Units at Close", OCC_UNITS, NUM,
          "102 Current + 8 Notice + 4 Evict. Case rules: 'Notice' and 'Evict' are treated as Occupied.",
          key="occunits")
r = a_val(r, "Vacant Units at Close", f"={A['units']}-{A['occunits']}", NUM,
          "All four are 'Vacant-Unrented'.", font=BLACK, key="vacunits")
r = a_val(r, "  In-Place Physical Occupancy", f"={A['occunits']}/{A['units']}", PCT, font=BLACK)
r = a_val(r, "In-Place Market Rent ($/month)", f"='Unit Mix'!{{mkt_tot}}", CUR, font=GREEN, key="mktrent")
r = a_val(r, "In-Place Actual Rent ($/month)", INPLACE_TOT, CUR,
          "Rent roll total row. Occupied units only.", key="actrent")
r = a_val(r, "  In-Place Loss-to-Lease (occupied units)", None, PCT, font=BLACK, key="ll_inplace")
r = a_val(r, "Units WITH In-Unit Washer/Dryer", N_WD, NUM,
          "Rent roll unit types containing 'w/ WD': Rhone 15, Roman 4, Cabernet 4, Brunswick 30.", key="n_wd")
r = a_val(r, "Units WITHOUT In-Unit Washer/Dryer", f"={A['units']}-{A['n_wd']}", NUM,
          "Abbey 15, Verdun 25, Bordeaux 25. Each carries the $5,000 addition per the case instructions.",
          font=BLACK, key="n_nowd")

r += 1
r = section(wsa, r, "RENOVATION PROGRAMME — TWO TRACKS")
w(wsa, f"A{r}", "The 53 units that already have laundry can be renovated IN PLACE (residents stay, no "
                "permit, no vacancy loss) but the premium is only captured at the next renewal. The 65 units "
                "that need laundry added require plumbing and electrical permits and a vacant unit, so they "
                "follow natural turnover — but the premium is captured immediately on the new lease.", NOTE,
  align="wrap")
wsa.merge_cells(f"A{r}:H{r}")
wsa.row_dimensions[r].height = 42
r += 1

w(wsa, f"A{r}", "TRACK A — 53 units with existing in-unit laundry (in-place)", BOLD, fill=FILL_SUB)
wsa.merge_cells(f"A{r}:H{r}")
r += 1
r = a_val(r, "  Units", f"={A['n_wd']}", NUM, "Rent roll unit types containing 'w/ WD'.", font=GREEN, key="a_units")
r = a_val(r, "  Start Month", 2, NUM,
          "No permit needed: cosmetic work (cabinets, counters, LVP, appliance and fixture swaps) is exempt "
          "in Marion County. Month 2 allows for GC mobilisation and a mock-up unit.", key="a_start")
r = a_val(r, "  Units per Month", 10, NUM,
          "In-place work does not consume a turn, so pace is limited by crew capacity rather than by lease "
          "expiries. 10/month across 5 buildings is one crew working 2-3 days per unit.", key="a_pace", fill=FILL_KEY)
r = a_val(r, "  Cost per Unit", 10000, CUR,
          "Lighter than a vacant-unit scope because floors under furniture and some millwork stay: cabinet "
          "refacing and paint, counters, appliances, lighting and plumbing fixtures, hardware.", key="a_cost")
r = a_val(r, "  Rent Premium ($/month)", 100, CUR,
          "Below the vacant-unit premium because the scope is lighter.", key="a_prem", fill=FILL_KEY)
r = a_val(r, "  Incremental Downtime (days)", 0, NUM, "Resident stays in occupancy throughout.", key="a_down")
r = a_val(r, "  Months to Capture Premium", 6, NUM,
          "You cannot raise an in-place resident's rent until renewal. Six months is the average time to the "
          "next lease expiry across a stabilised roll.", key="a_lag", fill=FILL_KEY)

r += 1
w(wsa, f"A{r}", "TRACK B — 65 units requiring laundry addition (vacant)", BOLD, fill=FILL_SUB)
wsa.merge_cells(f"A{r}:H{r}")
r += 1
r = a_val(r, "  Units", f"={A['n_nowd']}", NUM, "Abbey 15, Verdun 25, Bordeaux 25.", font=GREEN, key="b_units")
r = a_val(r, "  Start Month", 4, NUM,
          "Indiana permits multifamily in two sequential steps: a STATE Construction Design Release from IDHS "
          "(Class 1 structure; 10-business-day clock, 15-20 business days typical) BEFORE the local Marion "
          "County DBNS permit (5-20 days). Roughly 8-12 weeks from closing with design time.", key="b_start")
r = a_val(r, "  Units per Month", 5, NUM,
          "Constrained by natural turnover: T-12 physical occupancy of 93.2% implies roughly 55-60 turns a "
          "year. We do not force vacancy.", key="b_pace", fill=FILL_KEY)
r = a_val(r, "  Interior Cost per Unit", 13000, CUR,
          "Full vacant-unit scope: cabinets, counters, LVP throughout, appliances, lighting and plumbing "
          "fixtures, hardware, full paint.", key="b_cost_int")
r = a_val(r, "  W/D Addition Cost per Unit", 5000, CUR,
          "PRESCRIBED BY THE CASE: hookups plus the washer and dryer.", key="b_cost_wd")
r = a_val(r, "  Total Cost per Unit", None, CUR, font=BLACK, key="b_cost")
r = a_val(r, "  Rent Premium ($/month)", 200, CUR,
          "Full interior plus in-unit laundry. CBRE's own Upgrade Forecast puts $163 on the Verdun for a much "
          "lighter touch plus a washer/dryer.", key="b_prem", fill=FILL_KEY)
r = a_val(r, "  Incremental Downtime (days)", 14, NUM,
          "Over and above a normal turn. The renovation is executed inside the vacancy window.", key="b_down")
r = a_val(r, "  Months to Capture Premium", 1, NUM,
          "Captured immediately on the new lease.", key="b_lag")

r += 1
w(wsa, f"A{r}", "BLENDED PROGRAMME", BOLD, fill=FILL_SUB)
wsa.merge_cells(f"A{r}:H{r}")
r += 1
r = a_val(r, "  Blended Cost per Unit", None, CUR, font=BLACK, key="cost_blend")
r = a_val(r, "  Total Renovation Capital", None, CUR, font=BLACK, key="reno_total")
r = a_val(r, "  Blended Premium per Unit ($/month)", None, CUR, font=BLACK, key="prem_blend")
r = a_val(r, "  Return on Renovation Cost", None, PCT2, font=BLACK, key="roc")
r = a_val(r, "  Programme Completion (month of hold)", None, NUM, font=GREEN, key="done_month")
r = a_val(r, "Capex Contingency (% of all capital)", 0.08, PCT, key="conting")

r += 1
r = section(wsa, r, "REVENUE DRIVERS")
r = yr_hdr(r)
r = a_years(r, "Market Rent Growth", [0.015, 0.020, 0.025, 0.030, 0.030], PCT,
            "Yardi Matrix 1Q-2026: metro rents were FLAT on a trailing-three-month basis at $1,310 and up only "
            "1.1% year over year. Construction starts MORE THAN DOUBLED in 2025, so the delivery wave lands in "
            "2027-28, inside our hold. Year 1 is set slightly above the 1.1% actual, Year 2 held down by that "
            "wave, normalising to 3.0% only once it clears.", key="g_rent", fill=FILL_KEY)
r = a_years(r, "Loss to Lease (% of GPR)", [0.022, 0.020, 0.0175, 0.015, 0.015], PCT,
            "2.2% in place today. Burns down as the renovation programme resets leases to market.", key="ll")
r = a_years(r, "Operating Physical Vacancy", [0.055, 0.055, 0.055, 0.055, 0.055], PCT,
            "T-12 actual physical vacancy is 6.8%; CBRE forecasts 5.0% stabilised. Underwritten at 5.5% "
            "(94.5% occupancy) against a 94.0% metro figure: Carmel alone takes more than half of 2026 metro "
            "deliveries and sits immediately north of this asset across the county line. Renovation downtime is "
            "added separately from the Reno Schedule.", key="vac", fill=FILL_KEY)
r = a_years(r, "Concessions (% of GPR)", [0.0075, 0.0075, 0.005, 0.005, 0.005], PCT,
            "T-12 actual 0.5% and rising in recent months. Elevated while renovated units are first leased.",
            key="conc")
r = a_years(r, "Bad Debt (% of GPR)", [0.015, 0.0125, 0.0125, 0.010, 0.010], PCT,
            "T-12 net is 0.5% ($28,229 written off against $19,050 collected) but gross write-offs are "
            "accelerating and 4 units are in eviction with 8 on notice.", key="bd", fill=FILL_KEY)
r = a_years(r, "Other Income Growth", [0.03] * 5, PCT, key="g_oi")

r += 1
r = section(wsa, r, "OPERATING EXPENSES — YEAR 1 AND ANNUAL GROWTH")
w(wsa, f"A{r}", "", BOLD)
w(wsa, f"B{r}", "Year 1 $", BOLD, fill=FILL_SUB, align="center")
w(wsa, f"C{r}", "Growth", BOLD, fill=FILL_SUB, align="center")
w(wsa, f"H{r}", "Basis", BOLD, fill=FILL_SUB)
r += 1


def a_exp(row, label, y1, growth, note, key):
    w(wsa, f"A{row}", label, BLACK)
    w(wsa, f"B{row}", y1, BLUE, CUR, align="center")
    w(wsa, f"C{row}", growth, BLUE, PCT, align="center")
    w(wsa, f"H{row}", note, NOTE, align="wrap")
    A[key] = f"Assumptions!$B${row}"
    A[key + "_g"] = f"Assumptions!$C${row}"
    return row + 1


r = a_exp(r, "Payroll", 235000, 0.035,
          "T-12 $227,969; CBRE forecast $229,021 for 3.5 FTE. We add a second maintenance technician to "
          "execute 5 renovations a month.", "e_pay")
r = a_val(r, "Property Management Fee (% of EGR)", 0.04, PCT,
          "T-12 effective 3.5%; CBRE forecast 4.0%. Third-party market rate at this size.", key="e_mgmt")
r = a_exp(r, "Marketing & Advertising", 28000, 0.03,
          "T-12 $26,399. Increased for renovated-unit lease-up.", "e_mkt")
r = a_exp(r, "General & Administrative", 50000, 0.03,
          "T-12 $62,887 including $17,528 of legal and $14,730 of tax/audit concentrated in Dec-Jan (likely "
          "sale-related). CBRE forecast $42,209. Underwritten between the two.", "e_ga")
r = a_exp(r, "Bulk Internet (net expense)", 39577, 0.03,
          "$27.95/unit/month fully stabilised under the 7-year Comcast contract signed May 2025. T-12 shows "
          "only $11,068 because the contract is still ramping. Residents are billed $49 — see Other Income.", "e_int")
r = a_exp(r, "Electricity (common + vacant)", 31000, 0.03, "T-12 $30,320.", "e_elec")
r = a_exp(r, "Gas (central hot water + common)", 16000, 0.03, "T-12 $15,330.", "e_gas")
r = a_exp(r, "Water & Sewer", 90000, 0.04,
          "T-12 $86,421. 85% is recovered from residents — see Other Income.", "e_ws")
r = a_exp(r, "Repairs & Maintenance (incl. turns)", 255000, 0.03,
          "T-12 $250,935 ($2,127/unit) including $42,921 of turnover cost; CBRE forecast $247,656. No "
          "post-renovation savings credited.", "e_rm")
r = a_exp(r, "Insurance", 95000, 0.05,
          "T-12 $94,611 ($802/unit); CBRE forecast $88,326. No loss run or bindable quote provided.", "e_ins")
r = yr_hdr(r)
r = a_years(r, "Real Estate Taxes", [137000, 142500, 148200, 154100, 160300], CUR,
            "Marion County parcel 8000002 ACTUAL: 2025 net annual tax $129,648 on a $5,214,932 net assessed "
            "value at 2.4861% (gross AV $5,547,800 less a $332,868 '2% deduction'). The 25-year assessment "
            "record shows AV range-bound at $5.1-6.6mm since 2012 (0.35% a year) with the last assessment "
            "change on 03/15/2022, and the 2006 sale did NOT reset it. Base case grows the actual bill at 4% "
            "— above the 1.5% historical rate. A reassessment shock is a named sensitivity, not the base "
            "case. See the 'Tax Record' tab.", key="e_tax", fill=FILL_KEY)
r = a_val(r, "Replacement Reserves ($/unit/year)", 300, CUR,
          "Lender-standard for 1974 vintage. Below NOI. Grows with Other Income growth.", key="e_res")

r += 1
r = section(wsa, r, "DEBT")
r = a_val(r, "Maximum Loan to Cost", 0.65, PCT,
          "Value-add / bridge structure with a capital facility.", key="ltc", fill=FILL_KEY)
r = a_val(r, "Minimum Debt Yield (Year-1 NOI / loan)", 0.080, PCT,
          "Lender sizing test. Binding constraint at the base-case price — see Returns.", key="mindy")
r = a_val(r, "Minimum DSCR (Year-1)", 1.25, MULT,
          "Lender sizing test on an interest-only structure.", key="mindscr")
r = a_val(r, "Interest Rate (fixed, interest-only)", 0.0585, PCT2,
          "No terms provided. Approximates mid-2026 agency pricing on a 5-year interest-only structure.",
          key="rate", fill=FILL_KEY)
r = a_val(r, "Origination Fee (% of loan)", 0.01, PCT, key="fee")
r = a_val(r, "Other Financing Costs", 100000, CUR,
          "Lender legal, appraisal, engineering, cap-ex escrow set-up, rate-lock.", key="fincost")

# deferred formulas that need other sheets / rows
wsa[f"B{AROW['ppu']}"] = f"={A['price']}/{A['units']}"
wsa[f"B{AROW['ppsf']}"] = f"={A['price']}/{A['nrsf']}"
wsa[f"B{AROW['ll_inplace']}"] = (
    f"=1-{A['actrent']}/({A['mktrent']}-'Unit Mix'!{{vac_mkt}})")
wsa[f"B{AROW['vs_trade']}"] = f"={A['price']}/{A['actual']}-1"
wsa[f"B{AROW['b_cost']}"] = f"={A['b_cost_int']}+{A['b_cost_wd']}"
wsa[f"B{AROW['reno_total']}"] = f"={A['a_units']}*{A['a_cost']}+{A['b_units']}*{A['b_cost']}"
wsa[f"B{AROW['cost_blend']}"] = f"={A['reno_total']}/{A['units']}"
wsa[f"B{AROW['prem_blend']}"] = (
    f"=({A['a_units']}*{A['a_prem']}+{A['b_units']}*{A['b_prem']})/{A['units']}")
wsa[f"B{AROW['roc']}"] = f"={A['prem_blend']}*12/{A['cost_blend']}"

# ============================================================================
# 3. RENT ROLL
# ============================================================================
wsr = wb.create_sheet("Rent Roll")
wsr.sheet_view.showGridLines = False
title_block(wsr, "Rent Roll as of 07/01/2026 — with case status rules and per-unit renovation scope", "P")
hdrs = ["Unit", "BD/BA", "Unit Type", "SF", "RR Status", "UW Status", "Market Rent", "In-Place Rent",
        "Loss to Lease", "Lease From", "Lease To", "In-Unit W/D", "Reno Track", "Reno Cost $",
        "Premium $/mo", "Renovated Mkt Rent"]
widths = [10, 9, 17, 8, 15, 11, 12, 12, 11, 11, 11, 10, 18, 12, 12, 15]
for i, (h, wd) in enumerate(zip(hdrs, widths)):
    col = get_column_letter(i + 1)
    wsr.column_dimensions[col].width = wd
    w(wsr, f"{col}5", h, BOLD, fill=FILL_SUB, align="wrap", border=BOX)
wsr.row_dimensions[5].height = 30

r0 = 6
for i, x in enumerate(RR):
    rw = r0 + i
    occ = "vacant" not in x["status"].lower()
    haswd = x["utype"] in WD_TYPES
    w(wsr, f"A{rw}", x["unit"], BLACK)
    w(wsr, f"B{rw}", x["bdba"], BLACK, align="center")
    w(wsr, f"C{rw}", x["utype"], BLACK)
    w(wsr, f"D{rw}", x["sf"], BLACK, NUM, align="center")
    w(wsr, f"E{rw}", x["status"], BLACK)
    w(wsr, f"F{rw}", "Occupied" if occ else "Vacant", BOLD if not occ else BLACK, align="center")
    w(wsr, f"G{rw}", x["mkt"], BLACK, CUR)
    w(wsr, f"H{rw}", x["rent"] if occ else None, BLACK, CUR)
    w(wsr, f"I{rw}", f"=IF(F{rw}=\"Vacant\",0,H{rw}-G{rw})", BLACK, CUR)
    w(wsr, f"J{rw}", x["lf"] if occ else None, BLACK, DATE, align="center")
    w(wsr, f"K{rw}", x["lt"] if occ else None, BLACK, DATE, align="center")
    w(wsr, f"L{rw}", "Yes" if haswd else "No", BLACK, align="center")
    w(wsr, f"M{rw}", f"=IF(L{rw}=\"Yes\",\"A — in-place\",\"B — vacant + laundry\")", BLACK, align="center")
    w(wsr, f"N{rw}", f"=IF(L{rw}=\"Yes\",{A['a_cost']},{A['b_cost']})", BLACK, CUR)
    w(wsr, f"O{rw}", f"=IF(L{rw}=\"Yes\",{A['a_prem']},{A['b_prem']})", BLACK, CUR)
    w(wsr, f"P{rw}", f"=G{rw}+O{rw}", BLACK, CUR)
rN = r0 + len(RR) - 1
tot = rN + 1
w(wsr, f"A{tot}", f"Total / Average — {UNITS} Units", BOLD, border=TOP)
for col, f in [("D", f"=SUM(D{r0}:D{rN})"), ("G", f"=SUM(G{r0}:G{rN})"), ("H", f"=SUM(H{r0}:H{rN})"),
               ("I", f"=SUM(I{r0}:I{rN})"), ("N", f"=SUM(N{r0}:N{rN})"),
               ("O", f"=SUM(O{r0}:O{rN})"), ("P", f"=SUM(P{r0}:P{rN})")]:
    w(wsr, f"{col}{tot}", f, BOLD, CUR if col != "D" else NUM, border=TOP)
w(wsr, f"F{tot}", f"=COUNTIF(F{r0}:F{rN},\"Occupied\")&\" occ\"", BOLD, border=TOP, align="center")
w(wsr, f"L{tot}", f"=COUNTIF(L{r0}:L{rN},\"Yes\")&\" w/ W/D\"", BOLD, border=TOP, align="center")
wsr.merge_cells(f"A{tot+2}:P{tot+2}")
w(wsr, f"A{tot+2}",
  "Case status rules applied: 'Notice-Unrented' and 'Evict' are treated as Occupied; any variation of "
  "'Vacant' is treated as Vacant. Source: rent roll export dated 07/01/2026.", NOTE)
wsr.merge_cells(f"A{tot+3}:P{tot+3}")
w(wsr, f"A{tot+3}",
  "In-unit washer/dryer is taken from the rent roll unit-type description ('w/ WD'), which the OM's own "
  "floor-plan and building-mix schedules corroborate at 53 units. Per the data hierarchy the rent roll governs "
  "over the OM's conflicting '78 of 118' statement.", NOTE)
RR_RANGE = (r0, rN)

# ============================================================================
# 4. UNIT MIX
# ============================================================================
wsm = wb.create_sheet("Unit Mix")
wsm.sheet_view.showGridLines = False
title_block(wsm, "Floor Plan Mix, Renovation Scope and Renovated Rents (rent roll 07/01/2026)", "P")
cols = [("A", "Floor Plan", 18), ("B", "Type", 9), ("C", "Units", 8), ("D", "% of\nTotal", 8),
        ("E", "SF /\nUnit", 8), ("F", "Total SF", 10), ("G", "In-Unit\nW/D", 9),
        ("H", "Occ.\nUnits", 8), ("I", "Market\nRent", 11), ("J", "Market\nRent PSF", 11),
        ("K", "In-Place\nRent (occ)", 12), ("L", "Avg In-Place\nRent", 12),
        ("M", "Loss to\nLease", 10), ("N", "Reno Cost\n/ Unit", 11),
        ("O", "Reno Prem.\n/ Month", 11), ("P", "Renovated\nMkt Rent", 12),
        ("Q", "Renov.\nRent PSF", 11)]
for col, h, wd in cols:
    wsm.column_dimensions[col].width = wd
    w(wsm, f"{col}5", h, BOLD, fill=FILL_SUB, align="wrap", border=BOX)
wsm.row_dimensions[5].height = 32

m0 = 6
for i, ((bdba, utype, sf, mkt), d) in enumerate(MIX):
    rw = m0 + i
    haswd = utype in WD_TYPES
    w(wsm, f"A{rw}", utype, BLACK)
    w(wsm, f"B{rw}", bdba.replace(".00", "").replace(".50", ".5"), BLACK, align="center")
    w(wsm, f"C{rw}", d["n"], BLUE, NUM, align="center")
    w(wsm, f"D{rw}", f"=C{rw}/$C${m0+len(MIX)}", BLACK, PCT)
    w(wsm, f"E{rw}", sf, BLUE, NUM, align="center")
    w(wsm, f"F{rw}", f"=C{rw}*E{rw}", BLACK, NUM)
    w(wsm, f"G{rw}", "Yes" if haswd else "No", BLUE, align="center")
    w(wsm, f"H{rw}", d["occ"], BLUE, NUM, align="center")
    w(wsm, f"I{rw}", mkt, BLUE, CUR)
    w(wsm, f"J{rw}", f"=I{rw}/E{rw}", BLACK, CUR2)
    w(wsm, f"K{rw}", d["rent"], BLUE, CUR)
    w(wsm, f"L{rw}", f"=IF(H{rw}=0,0,K{rw}/H{rw})", BLACK, CUR)
    w(wsm, f"M{rw}", f"=IF(H{rw}=0,0,L{rw}/I{rw}-1)", BLACK, PCT)
    w(wsm, f"N{rw}", f"=IF(G{rw}=\"Yes\",{A['a_cost']},{A['b_cost']})", BLACK, CUR)
    w(wsm, f"O{rw}", f"=IF(G{rw}=\"Yes\",{A['a_prem']},{A['b_prem']})", BLACK, CUR)
    w(wsm, f"P{rw}", f"=I{rw}+O{rw}", BOLD, CUR)
    w(wsm, f"Q{rw}", f"=P{rw}/E{rw}", BLACK, CUR2)
mN = m0 + len(MIX) - 1
mt = mN + 1
w(wsm, f"A{mt}", "Total / Weighted Average", BOLD, border=TOP)
for col, f, fmt in [
    ("C", f"=SUM(C{m0}:C{mN})", NUM), ("D", f"=SUM(D{m0}:D{mN})", PCT),
    ("E", f"=F{mt}/C{mt}", NUM), ("F", f"=SUM(F{m0}:F{mN})", NUM),
    ("H", f"=SUM(H{m0}:H{mN})", NUM),
    ("I", f"=SUMPRODUCT(C{m0}:C{mN},I{m0}:I{mN})/C{mt}", CUR),
    ("J", f"=SUMPRODUCT(C{m0}:C{mN},I{m0}:I{mN})/F{mt}", CUR2),
    ("K", f"=SUM(K{m0}:K{mN})", CUR), ("L", f"=K{mt}/H{mt}", CUR),
    ("M", f"=L{mt}/(SUMPRODUCT(H{m0}:H{mN},I{m0}:I{mN})/H{mt})-1", PCT),
    ("N", f"=SUMPRODUCT(C{m0}:C{mN},N{m0}:N{mN})/C{mt}", CUR),
    ("O", f"=SUMPRODUCT(C{m0}:C{mN},O{m0}:O{mN})/C{mt}", CUR),
    ("P", f"=SUMPRODUCT(C{m0}:C{mN},P{m0}:P{mN})/C{mt}", CUR),
    ("Q", f"=SUMPRODUCT(C{m0}:C{mN},P{m0}:P{mN})/F{mt}", CUR2),
]:
    w(wsm, f"{col}{mt}", f, BOLD, fmt, fill=FILL_TOT, border=TOP)
w(wsm, f"G{mt}", f"=SUMIF(G{m0}:G{mN},\"Yes\",C{m0}:C{mN})&\" of \"&C{mt}", BOLD, border=TOP, align="center")

# monthly / annual roll-up block used by the model
b = mt + 2
wsm.merge_cells(f"A{b}:D{b}")
w(wsm, f"A{b}", "AGGREGATE RENT POTENTIAL", BOLD, fill=FILL_SUB)
rows_agg = [
    ("Total Market Rent — Classic Condition ($/month)", f"=SUMPRODUCT(C{m0}:C{mN},I{m0}:I{mN})", CUR, "mkt_mo"),
    ("Total Market Rent — Renovated ($/month)", f"=SUMPRODUCT(C{m0}:C{mN},P{m0}:P{mN})", CUR, "reno_mo"),
    ("Average Market Rent / Unit — Classic", f"=D{b+1}/C{mt}", CUR, "mkt_avg"),
    ("Average Market Rent / Unit — Renovated", f"=D{b+2}/C{mt}", CUR, "reno_avg"),
    ("Market Rent of the 4 Vacant Units ($/month)",
     f"=SUMPRODUCT((C{m0}:C{mN}-H{m0}:H{mN}),I{m0}:I{mN})", CUR, "vac_mkt"),
    ("Total Renovation Capital ($)", f"=SUMPRODUCT(C{m0}:C{mN},N{m0}:N{mN})", CUR, "reno_cap"),
    ("Units Requiring W/D Addition", f"=SUMIF(G{m0}:G{mN},\"No\",C{m0}:C{mN})", NUM, "n_nowd"),
]
MIXREF = {}
for i, (lbl, f, fmt, key) in enumerate(rows_agg):
    rw = b + 1 + i
    wsm.merge_cells(f"A{rw}:C{rw}")
    w(wsm, f"A{rw}", lbl, BLACK)
    w(wsm, f"D{rw}", f, BOLD, fmt)
    MIXREF[key] = f"'Unit Mix'!$D${rw}"
wsm.merge_cells(f"A{b+len(rows_agg)+2}:Q{b+len(rows_agg)+2}")
w(wsm, f"A{b+len(rows_agg)+2}",
  "Unit counts, square footage, market rents and in-place rents are taken directly from the 07/01/2026 rent "
  "roll and foot to its total row ($153,435 market / $144,804 in-place / 134,371 SF). Renovation cost and "
  "premium are driven from the Assumptions tab.", NOTE)

# patch Assumptions formulas that referenced Unit Mix
wsa[f"B{AROW['mktrent']}"] = f"={MIXREF['mkt_mo']}"
wsa[f"B{AROW['ll_inplace']}"] = f"=1-{A['actrent']}/({A['mktrent']}-{MIXREF['vac_mkt']})"

# ============================================================================
# 5. T-12 RECAST
# ============================================================================
wst = wb.create_sheet("T-12 Recast")
wst.sheet_view.showGridLines = False
wst.column_dimensions["A"].width = 42
for c in ["B", "C", "D", "E"]:
    wst.column_dimensions[c].width = 14
wst.column_dimensions["F"].width = 76
title_block(wst, "T-12 Actuals (Feb 2025 – Jan 2026) Recast to Underwriting Line Items", "F")


def g(*labels):
    """Sum of T-12 account totals."""
    s = 0.0
    for lab in labels:
        for k, v in T12.items():
            if k.startswith(lab):
                s += v
                break
        else:
            raise KeyError(lab)
    return s


ADV = g("6210", "6211", "6213", "6214", "6215", "6216", "6217", "6220")
GA = g("6275", "6311", "6312", "6313", "6314", "6315", "6317", "6318", "6341",
       "6342", "6343", "6344", "6345", "6346", "6380", "6381", "6391", "6392", "6393")
INTERNET = g("6316")
MGMT = g("6320")
ELEC = g("6449", "6450")
GAS = g("6452")
WS = g("6451")
UTIL_OTH = g("6454")
LABOR = g("6510", "6511")
MAINT = g("6509 Total Maintenance Expense")
TAX = g("6710")
INS = g("6720")
GPR_T12 = g("5120")
ECON_VAC = g("5210")
PHYS_VAC = g("5220")
CONC = g("5250")
BD_NET = g("5925") + g("5926")
OI_LINES = [
    ("Carport / Garage Rent (74 carports @ $25)", g("5170")),
    ("Storage Rent (5 of 23 rented @ $15-20)", g("5171")),
    ("Credit Builder Fees", g("5175")),
    ("Onsite Transfer Fees", g("5181")),
    ("Administrative Fees", g("5182")),
    ("Utility Fees (water/sewer + valet trash bundled)", g("5185")),
    ("Technology Fee ($49 bulk internet, ramping)", g("5187")),
    ("Lease Termination Fees", g("5188")),
    ("Eviction Fees", g("5189")),
    ("Washer/Dryer Rental (programme discontinued)", g("5191")),
    ("Application Fees", g("5192")),
    ("Utility Reimbursement", g("5193")),
    ("Pet Fees", g("5194")),
    ("Short-Term Lease Fees", g("5195")),
    ("Late Fees", g("5196")),
    ("NSF Fees", g("5197")),
    ("Miscellaneous Income", g("5199")),
    ("Damage Charges (cleaning, flooring, paint, other)",
     g("5930") + g("5931") + g("5932") + g("5933")),
]
OI_OPER = sum(v for _, v in OI_LINES)
INTEREST_INC = g("5410")

r = 5
r = section(wst, r, "REVENUE", "F")
w(wst, f"A{r}", "", BOLD)
for col, h in [("B", "T-12 Actual"), ("C", "Per Unit"), ("D", "UW Year 1"), ("E", "Variance")]:
    w(wst, f"{col}{r}", h, BOLD, fill=FILL_SUB, align="center")
w(wst, f"F{r}", "Underwriting note", BOLD, fill=FILL_SUB)
r += 1
TREF = {}


def t_row(row, label, t12v, uw_formula, note, bold=False, border=None, key=None):
    w(wst, f"A{row}", label, BOLD if bold else BLACK, border=border)
    w(wst, f"B{row}", t12v, BLUE if not bold else BOLD, CUR, border=border)
    w(wst, f"C{row}", f"=B{row}/{A['units']}" if t12v is not None else None,
      BLACK if not bold else BOLD, CUR, border=border)
    if uw_formula is not None:
        w(wst, f"D{row}", uw_formula, GREEN if not bold else BOLD_GREEN, CUR, border=border)
        w(wst, f"E{row}", f"=IF(B{row}=0,\"n/a\",D{row}/B{row}-1)", BLACK if not bold else BOLD,
          PCT, border=border)
    w(wst, f"F{row}", note, NOTE, align="wrap")
    if key:
        TREF[key] = f"'T-12 Recast'!$B${row}"
    return row + 1


AM = "'Annual Model'"
r = t_row(r, "Gross Rent Potential", GPR_T12, f"={AM}!C{{gpr}}",
          "T-12 gross rent potential at then-current market rents. UW Year 1 reflects the 07/01/2026 rent roll, "
          "renovated-unit premiums as they come online, and market rent growth.", key="gpr")
r = t_row(r, "Economic Vacancy / Loss to Lease", ECON_VAC, f"={AM}!C{{ll}}",
          "T-12 loss to lease is 1.7% of GPR. Rent roll shows 2.2% in place today.")
r = t_row(r, "Physical Vacancy", PHYS_VAC, f"={AM}!C{{vac}}+{AM}!C{{dvac}}",
          "T-12 physical vacancy is 9.4% of GPR (and deteriorating: Jan-26 alone was $20,757). UW is 5.0% "
          "operating plus renovation downtime.")
r = t_row(r, "Concessions", CONC, f"={AM}!C{{conc}}",
          "T-12 $9,688. Concessions restarted in Aug-25 after a year at zero.")
r = t_row(r, "Bad Debt, net of collections", BD_NET, f"={AM}!C{{bd}}",
          f"T-12 gross write-offs $28,229 less $19,050 collected = ${abs(BD_NET):,.0f} net (0.5%). "
          "Underwritten at 1.5% falling to 1.0% given 4 evictions and 8 notices on the rent roll.")
r = t_row(r, "Net Rental Income", GPR_T12 + ECON_VAC + PHYS_VAC + CONC + BD_NET, f"={AM}!C{{nri}}",
          "", bold=True, border=TOP)
T12_NRI_ROW = r - 1
r += 1

w(wst, f"A{r}", "OTHER INCOME", BOLD, fill=FILL_SUB)
r += 1
oi0 = r
for lbl, v in OI_LINES:
    w(wst, f"A{r}", "   " + lbl, BLACK)
    w(wst, f"B{r}", v, BLUE, CUR)
    w(wst, f"C{r}", f"=B{r}/{A['units']}", BLACK, CUR)
    r += 1
oiN = r - 1
r = t_row(r, "Total Operating Other Income", f"=SUM(B{oi0}:B{oiN})", f"='Other Income'!C{{oitot}}",
          "Interest income of $4,133 is excluded as non-property income. Bad debt is shown above as a revenue "
          "deduction rather than netted in other income.", bold=True, border=TOP)
wst[f"B{r-1}"].font = BOLD
T12_OI_ROW = r - 1
r = t_row(r, "Memo: Interest Income (excluded)", INTEREST_INC, None,
          "Excluded from underwritten revenue — a function of the seller's cash balances, not the real estate.")
r += 1

r = section(wst, r, "OPERATING EXPENSES", "F")
w(wst, f"A{r}", "", BOLD)
for col, h in [("B", "T-12 Actual"), ("C", "Per Unit"), ("D", "UW Year 1"), ("E", "Variance")]:
    w(wst, f"{col}{r}", h, BOLD, fill=FILL_SUB, align="center")
w(wst, f"F{r}", "Underwriting note", BOLD, fill=FILL_SUB)
r += 1
exp_rows = []
for label, t12v, uwref, note in [
    ("Payroll (incl. contract labour)", LABOR, A["e_pay"],
     "T-12 $227,969 ($1,932/unit) including $9,778 of contract labour. CBRE forecast $229,021 for a manager, "
     "part-time leasing, maintenance supervisor and one technician. We fund a second technician."),
    ("Property Management Fee", MGMT, f"{AM}!C{{mgmt}}",
     "T-12 is an effective 3.5% of revenue. Underwritten at a market 4.0% of effective gross revenue."),
    ("Marketing & Advertising", ADV, A["e_mkt"],
     "T-12 $26,399, of which $22,047 is ILS. Increased modestly for renovated-unit lease-up."),
    ("General & Administrative", GA, A["e_ga"],
     "T-12 $62,887 includes $17,528 legal and $14,730 tax/audit concentrated in Dec-25/Jan-26, almost certainly "
     "transaction-related. CBRE forecast $42,209. Underwritten at $50,000 between the two."),
    ("Bulk Internet", INTERNET, A["e_int"],
     "T-12 $11,068 captures only 7 months of a ramping contract. Stabilised is $27.95/unit/month = $39,577. "
     "Understating this line is the easiest way to overstate NOI on this deal."),
    ("Electricity (common + vacant)", ELEC, A["e_elec"], "T-12 $30,320 across corporate and vacant-unit meters."),
    ("Gas (central hot water + common)", GAS, A["e_gas"], "T-12 $15,330. Central gas water heaters plus 8 individual electric."),
    ("Water & Sewer", WS + UTIL_OTH, A["e_ws"],
     "T-12 $86,468. Master-metered and billed back to residents; recovery is modelled at 85% in Other Income."),
    ("Repairs & Maintenance (incl. turns)", MAINT, A["e_rm"],
     "T-12 $250,935 ($2,127/unit) including $42,921 of turnover cost. No post-renovation savings credited, "
     "though a fully renovated portfolio should run lighter."),
    ("Insurance", INS, A["e_ins"],
     "T-12 $94,611 ($802/unit). CBRE forecast $88,326. Held at T-12 plus a small cushion pending a loss run."),
    ("Real Estate Taxes", TAX, f"{AM}!C{{tax}}",
     "T-12 $149,677 including a $17,958 October catch-up. CBRE's 2025-pay-2026 estimate with the SB-1 deduction "
     "is $135,374. Year 1 $140,000 stepping to $210,000 by Year 5 for reassessment on sale."),
]:
    w(wst, f"A{r}", label, BLACK)
    w(wst, f"B{r}", t12v, BLUE, CUR)
    w(wst, f"C{r}", f"=B{r}/{A['units']}", BLACK, CUR)
    w(wst, f"D{r}", f"={uwref}", GREEN, CUR)
    w(wst, f"E{r}", f"=D{r}/B{r}-1", BLACK, PCT)
    w(wst, f"F{r}", note, NOTE, align="wrap")
    wst.row_dimensions[r].height = 30
    exp_rows.append(r)
    r += 1
e0, eN = exp_rows[0], exp_rows[-1]
w(wst, f"A{r}", "Total Operating Expenses", BOLD, border=TOP)
w(wst, f"B{r}", f"=SUM(B{e0}:B{eN})", BOLD, CUR, border=TOP)
w(wst, f"C{r}", f"=B{r}/{A['units']}", BOLD, CUR, border=TOP)
w(wst, f"D{r}", f"={AM}!C{{opex}}", BOLD_GREEN, CUR, border=TOP)
w(wst, f"E{r}", f"=D{r}/B{r}-1", BOLD, PCT, border=TOP)
T12_OPEX_ROW = r
r += 1
w(wst, f"A{r}", "Net Operating Income", BOLD, border=TOPBOT)
w(wst, f"B{r}", f"=B{T12_NRI_ROW}+B{T12_OI_ROW}-B{T12_OPEX_ROW}", BOLD, CUR, border=TOPBOT)
w(wst, f"C{r}", f"=B{r}/{A['units']}", BOLD, CUR, border=TOPBOT)
w(wst, f"D{r}", f"={AM}!C{{noi}}", BOLD_GREEN, CUR, border=TOPBOT)
w(wst, f"E{r}", f"=D{r}/B{r}-1", BOLD, PCT, border=TOPBOT)
r += 2
w(wst, f"A{r}", "Memo — items the property books below NOI or excludes:", BOLD)
r += 1
for lbl, v, note in [
    ("Replacement Reserve / capital spend (T-12 actual)", g("6799 Total Replacement Reserve"),
     "The property expensed $108,412 of capital below the line in the T-12, including $34,535 of "
     "concrete/asphalt and $29,053 of water heaters. Evidence of real ongoing capital need."),
    ("Depreciation & Amortisation", g("6800 Total Depreciation & Amortization"), "Non-cash; excluded."),
]:
    w(wst, f"A{r}", "   " + lbl, BLACK)
    w(wst, f"B{r}", v, BLUE, CUR)
    w(wst, f"C{r}", f"=B{r}/{A['units']}", BLACK, CUR)
    w(wst, f"F{r}", note, NOTE, align="wrap")
    r += 1
w(wst, f"A{r+1}",
  "Source: Chateau in the Woods Trailing Profit And Loss Detail, January 2026, accrual basis (ResMan), "
  "covering February 2025 through January 2026. All T-12 figures are the report's own Adjusted Total column.",
  NOTE)
w(wst, f"A{r+2}",
  "Tie-out: the property's own statement shows total income of $1,790,101 less total expense of $1,017,727 = "
  "$772,374 of NOI. This recast shows $768,242, the difference being the $4,133 of interest income excluded "
  "above as non-property income. Bad debt is presented here as a revenue deduction rather than netted inside "
  "other income, which changes the presentation but not the NOI.", NOTE)
T12_NRI = GPR_T12 + ECON_VAC + PHYS_VAC + CONC + BD_NET

# ============================================================================
# 6. OTHER INCOME
# ============================================================================
wso = wb.create_sheet("Other Income")
wso.sheet_view.showGridLines = False
wso.column_dimensions["A"].width = 44
wso.column_dimensions["B"].width = 13
for c in YR_COLS:
    wso.column_dimensions[c].width = 13
wso.column_dimensions["H"].width = 62
title_block(wso, "Other Income Build — T-12 Actuals plus Identified Ancillary Upside")

r = 5
w(wso, f"A{r}", "", BOLD)
w(wso, f"B{r}", "T-12 Actual", BOLD, fill=FILL_SUB, align="center")
for i, c in enumerate(YR_COLS):
    w(wso, f"{c}{r}", f"Year {i+1}", BOLD, fill=FILL_SUB, align="center")
w(wso, f"H{r}", "Basis", BOLD, fill=FILL_SUB)
r += 1
OIR = {}

# implementation ramp inputs
ramp_start = r
for lbl, vals, key, note in [
    ("Ramp — Technology Fee ($49) implementation", [0.85, 1.0, 1.0, 1.0, 1.0], "ramp_tech",
     "Contract signed May 2025 and implemented on new leases and renewals over 13 months. Jan-26 booked "
     "$3,303 of a $5,489 full month, i.e. 60% penetrated at the T-12 close."),
    ("Ramp — Valet Trash ($25) implementation", [0.75, 0.95, 1.0, 1.0, 1.0], "ramp_valet",
     "Charged on new leases and renewals. The property begins paying the valet contractor in Jan-2026."),
    ("Ramp — Pet Rent (new policy)", [0.50, 1.0, 1.0, 1.0, 1.0], "ramp_pet",
     "The property charges no pet fee or pet rent today. New policy phased in on renewals over year one."),
]:
    w(wso, f"A{r}", lbl, BLACK)
    for i, c in enumerate(YR_COLS):
        w(wso, f"{c}{r}", vals[i], BLUE, PCT, align="center")
    w(wso, f"H{r}", note, NOTE, align="wrap")
    wso.row_dimensions[r].height = 28
    OIR[key] = r
    r += 1

w(wso, f"A{r}", "Economic occupancy factor for per-unit charges", BLACK)
for c in YR_COLS:
    w(wso, f"{c}{r}", 0.95, BLUE, PCT, align="center")
OIR["occf"] = r
w(wso, f"H{r}", "Consistent with CBRE's other-income forecast method.", NOTE, align="wrap")
r += 2

w(wso, f"A{r}", "OTHER INCOME LINE ITEMS", BOLD, fill=FILL_SUB)
r += 1
oib = r
lines = []


def oi_line(row, label, t12v, y1_formula, note, growth=True):
    w(wso, f"A{row}", label, BLACK)
    w(wso, f"B{row}", t12v, BLUE, CUR)
    w(wso, f"C{row}", y1_formula, BLACK, CUR)
    for i, c in enumerate(YR_COLS[1:], start=1):
        prev = YR_COLS[i - 1]
        if growth is True:
            w(wso, f"{c}{row}", f"={prev}{row}*(1+Assumptions!{c}${AROW['g_oi']})", BLACK, CUR)
        elif growth == "recover":
            w(wso, f"{c}{row}", None, BLACK, CUR)   # filled after
        elif growth == "ramp_tech":
            w(wso, f"{c}{row}", f"={prev}{row}/{prev}${OIR['ramp_tech']}*{c}${OIR['ramp_tech']}"
                                f"*(1+Assumptions!{c}${AROW['g_oi']})", BLACK, CUR)
        elif growth == "ramp_valet":
            w(wso, f"{c}{row}", f"={prev}{row}/{prev}${OIR['ramp_valet']}*{c}${OIR['ramp_valet']}"
                                f"*(1+Assumptions!{c}${AROW['g_oi']})", BLACK, CUR)
        elif growth == "ramp_pet":
            w(wso, f"{c}{row}", f"={prev}{row}/{prev}${OIR['ramp_pet']}*{c}${OIR['ramp_pet']}"
                                f"*(1+Assumptions!{c}${AROW['g_oi']})", BLACK, CUR)
    w(wso, f"H{row}", note, NOTE, align="wrap")
    wso.row_dimensions[row].height = 26
    lines.append(row)
    return row + 1


d = dict(OI_LINES)
r = oi_line(r, "Water & Sewer Recovery (85% of expense)", d["Utility Fees (water/sewer + valet trash bundled)"],
            f"=0.85*Assumptions!$B${openpyxl.utils.cell.coordinate_from_string(A['e_ws'].split('!')[1])[1]}",
            "T-12 Utility Fees of $97,812 bundle water/sewer and valet trash. CBRE's forecast splits them and "
            "assumes 85% water/sewer recovery. Flexes with the water & sewer expense line.", growth="recover")
WS_REC_ROW = r - 1
r = oi_line(r, "Valet Trash ($25/unit/month)", 0,
            f"=25*12*{A['units']}*C${OIR['occf']}*C${OIR['ramp_valet']}",
            "Broken out of Utility Fees per CBRE. A valet contract has been signed and the property starts "
            "paying for the service in Jan-2026, so the charge must be collected.", growth="ramp_valet")
r = oi_line(r, "Technology Fee ($49/unit/month)", d["Technology Fee ($49 bulk internet, ramping)"],
            f"=49*12*{A['units']}*C${OIR['occf']}*C${OIR['ramp_tech']}",
            "Offsets the $27.95/unit/month bulk internet expense. Contractual and already being billed.",
            growth="ramp_tech")
r = oi_line(r, "Carport Rent (74 carports @ $25/month)", d["Carport / Garage Rent (74 carports @ $25)"],
            f"=B{r}", "T-12 actual. 74 carports at $25; occupancy of the carports, not the units.")
r = oi_line(r, "Pet Rent ($30/month on 25% of units)", d["Pet Fees"],
            f"=30*12*{A['units']}*0.25*C${OIR['ramp_pet']}",
            "The property currently charges nothing for pets. A conservative $30/month on 25% of units; CBRE "
            "forecasts the same $10,620.", growth="ramp_pet")
r = oi_line(r, "Storage Rent (23 units available)", d["Storage Rent (5 of 23 rented @ $15-20)"], "=3600",
            "Only 5 of 23 storage units are rented today at $15-20. Underwritten at 12 rented at $25.")
r = oi_line(r, "Application & Administrative Fees",
            d["Application Fees"] + d["Administrative Fees"] + d["Credit Builder Fees"] + d["Onsite Transfer Fees"],
            f"=B{r}", "T-12 actual: application, administrative, credit builder and transfer fees.")
r = oi_line(r, "Late & NSF Fees", d["Late Fees"] + d["NSF Fees"], f"=B{r}", "T-12 actual.")
r = oi_line(r, "Short-Term Lease Premiums", d["Short-Term Lease Fees"], f"=B{r}",
            "T-12 actual. $75/month on 6-8 month renewals, $50/month on 9-11 month renewals.")
r = oi_line(r, "Lease Termination & Eviction Fees",
            d["Lease Termination Fees"] + d["Eviction Fees"], "=4000",
            f"T-12 ${d['Lease Termination Fees'] + d['Eviction Fees']:,.0f}. Haircut — volatile and partly a "
            "function of the current resident base.")
r = oi_line(r, "Damage Charges Recovered (net of SODA)",
            d["Damage Charges (cleaning, flooring, paint, other)"], "=15000",
            f"T-12 ${d['Damage Charges (cleaning, flooring, paint, other)']:,.0f}. Haircut; should fall as the "
            "resident profile improves post-renovation.")
r = oi_line(r, "Miscellaneous & Utility Reimbursement",
            d["Miscellaneous Income"] + d["Utility Reimbursement"], "=5000", "T-12 actual, lightly haircut.")
r = oi_line(r, "Washer/Dryer Rental", d["Washer/Dryer Rental (programme discontinued)"], "=0",
            "Programme discontinued, and in-unit laundry is now monetised through the rent premium. Set to zero "
            "to avoid double-counting.")

# water & sewer recovery growth columns tie to the W/S expense growth
e_ws_row = int(A["e_ws"].split("$")[-1])
for i, c in enumerate(YR_COLS[1:], start=1):
    prev = YR_COLS[i - 1]
    wso[f"{c}{WS_REC_ROW}"] = f"={prev}{WS_REC_ROW}*(1+Assumptions!$C${e_ws_row})"
    wso[f"{c}{WS_REC_ROW}"].font = BLACK
    wso[f"{c}{WS_REC_ROW}"].number_format = CUR

l0, lN = lines[0], lines[-1]
w(wso, f"A{r}", "Total Other Income", BOLD, border=TOP)
w(wso, f"B{r}", f"=SUM(B{l0}:B{lN})", BOLD, CUR, border=TOP)
for c in YR_COLS:
    w(wso, f"{c}{r}", f"=SUM({c}{l0}:{c}{lN})", BOLD, CUR, border=TOP)
OI_TOT_ROW = r
r += 1
w(wso, f"A{r}", "   Per Unit", BLACK)
w(wso, f"B{r}", f"=B{OI_TOT_ROW}/{A['units']}", BLACK, CUR)
for c in YR_COLS:
    w(wso, f"{c}{r}", f"={c}{OI_TOT_ROW}/{A['units']}", BLACK, CUR)
r += 1
w(wso, f"A{r}", "   % of Effective Gross Revenue", BLACK)
for c in YR_COLS:
    w(wso, f"{c}{r}", f"={c}{OI_TOT_ROW}/'Annual Model'!{c}{{egr}}", BLACK, PCT)
OI_PCT_ROW = r
r += 2
wso.merge_cells(f"A{r}:H{r}")
w(wso, f"A{r}",
  f"T-12 total operating other income of ${OI_OPER:,.0f} (${OI_OPER/UNITS:,.0f}/unit) excludes $4,133 of "
  "interest income and treats bad debt as a revenue deduction. CBRE's Acquisition Forecast is $263,803 "
  "($2,236/unit); the ancillary upside underwritten here — fully implementing the technology fee and valet "
  "trash charge, and introducing pet rent — is the same upside the OM identifies, phased in over the ramps "
  "shown above rather than assumed on day one.", NOTE, align="wrap")
wso.row_dimensions[r].height = 60

# ============================================================================
# 7. CAPEX BUDGET
# ============================================================================
wsc = wb.create_sheet("Capex Budget")
wsc.sheet_view.showGridLines = False
wsc.column_dimensions["A"].width = 46
wsc.column_dimensions["B"].width = 13
wsc.column_dimensions["C"].width = 12
for c in ["D", "E", "F", "G", "H"]:
    wsc.column_dimensions[c].width = 11
wsc.column_dimensions["I"].width = 64
title_block(wsc, "Non-Unit Capital Budget — Deferred Maintenance, Building Systems & Amenities", "I")

r = 5
w(wsc, f"A{r}", "Item", BOLD, fill=FILL_SUB)
w(wsc, f"B{r}", "Total Cost", BOLD, fill=FILL_SUB, align="center")
w(wsc, f"C{r}", "Per Unit", BOLD, fill=FILL_SUB, align="center")
for i, c in enumerate(["D", "E", "F", "G", "H"]):
    w(wsc, f"{c}{r}", f"Yr {i+1} %", BOLD, fill=FILL_SUB, align="center")
w(wsc, f"I{r}", "Scope and source", BOLD, fill=FILL_SUB)
r += 1
cap0 = r
CAPEX_ITEMS = [
    ("Roof replacement — 3.5 of 5 roofs", 360000, [0, .40, .40, .20, 0],
     "OM: '3.5 of 5 roofs are older than 2006. They are rubber roofs. Assume they will need to be replaced "
     "within 10 years.' Building 4 cost $90,000 in 2020 and half of Building 5 cost $54,507 in 2024. A buyer "
     "in five years will price any roof we do not replace, so it is underwritten inside the hold."),
    ("Asphalt patch, sealcoat & stripe", 150000, [1, 0, 0, 0, 0],
     "OM immediate need: parking lot and carport areas. The property spent $34,535 on concrete/asphalt in the "
     "T-12 alone and $38,749 in 2022 — patch-and-repair, not a programme."),
    ("Sump pumps (4) & patio drain cleanout", 40000, [1, 0, 0, 0, 0],
     "OM immediate need: 37 below-grade units, 4 sump pumps needed. Water intrusion risk in below-grade units "
     "is a live liability, not a cosmetic item."),
    ("Window & sliding patio door glass", 150000, [.34, .33, .33, 0, 0],
     "OM: original aluminium sliders and sliding patio doors, fogged glass replaced on turns. Executed with the "
     "interior renovation on each turn."),
    ("Building entry system upgrades (20 entries)", 60000, [.50, .50, 0, 0, 0],
     "OM value-add item: upgrade building entry systems across the 20 building entries."),
    ("Clubhouse, business centre & fitness refresh", 75000, [1, 0, 0, 0, 0],
     "OM suggested amenity: update clubhouse furniture and decor. Required to support renovated rents."),
    ("Dog park, grilling stations & pool deck seating", 50000, [1, 0, 0, 0, 0],
     "OM suggested amenities. Low cost, high visibility on tours; supports the rent premium."),
    ("Exterior paint, landscaping & signage refresh", 100000, [.50, .50, 0, 0, 0],
     "Curb appeal. Siding is fibre cement and was replaced in 2020 ($225,000), so this is paint and grounds, "
     "not envelope."),
]
for lbl, cost, ph, note in CAPEX_ITEMS:
    w(wsc, f"A{r}", lbl, BLACK)
    w(wsc, f"B{r}", cost, BLUE, CUR)
    w(wsc, f"C{r}", f"=B{r}/{A['units']}", BLACK, CUR)
    for i, c in enumerate(["D", "E", "F", "G", "H"]):
        w(wsc, f"{c}{r}", ph[i], BLUE, PCT, align="center")
    w(wsc, f"I{r}", note, NOTE, align="wrap")
    wsc.row_dimensions[r].height = 40
    r += 1
capN = r - 1
w(wsc, f"A{r}", "Total Non-Unit Capital", BOLD, border=TOP)
w(wsc, f"B{r}", f"=SUM(B{cap0}:B{capN})", BOLD, CUR, border=TOP)
w(wsc, f"C{r}", f"=B{r}/{A['units']}", BOLD, CUR, border=TOP)
for c in ["D", "E", "F", "G", "H"]:
    w(wsc, f"{c}{r}", f"=SUMPRODUCT($B${cap0}:$B${capN},{c}{cap0}:{c}{capN})", BOLD, CUR, border=TOP)
CAP_TOT_ROW = r
r += 1
w(wsc, f"A{r}", "   Check: phasing sums to 100%", ITAL)
w(wsc, f"B{r}", f"=SUM(D{r-1}:H{r-1})/B{r-1}", BLACK, PCT)
r += 2
wsc.merge_cells(f"A{r}:I{r}")
w(wsc, f"A{r}",
  "Excluded from this budget by choice, and therefore from the returns: converting the Building 4042 "
  "maintenance storage into two additional apartments, and converting the back 9 carports to garages plus a "
  "new maintenance shop. Both are identified in the OM and both are genuine, but at roughly $310,000 and "
  "$180,000 respectively they return only about 5-10% on cost against a 12.7% interior programme. They are "
  "upside we would pursue with a second look at scope and pricing, not underwritten value.", NOTE, align="wrap")
wsc.row_dimensions[r].height = 60

# ============================================================================
# 8. RENO SCHEDULE
# ============================================================================
wsn = wb.create_sheet("Reno Schedule")
wsn.sheet_view.showGridLines = False
title_block(wsn, "60-Month Two-Track Renovation Schedule — Delivery, Premium Capture, Capital and Downtime", "P")
nh = [("A", "Month", 8), ("B", "Date", 10), ("C", "Model\nYear", 8),
      ("D", "A: Units\nDone", 10), ("E", "A: Cum.\nDone", 10), ("F", "A: Units\nEarning", 11),
      ("G", "B: Units\nDone", 10), ("H", "B: Cum.\nDone", 10), ("I", "B: Units\nEarning", 11),
      ("J", "Total\nRenovated", 11), ("K", "Units Earning\nPremium", 12),
      ("L", "Premium in\nForce ($/mo)", 13), ("M", "Track A\nCapex", 11),
      ("N", "Track B\nCapex", 11), ("O", "Total\nCapex", 12),
      ("P", "Downtime Vac.\n(% of GPR)", 13)]
for col, h, wd in nh:
    wsn.column_dimensions[col].width = wd
    w(wsn, f"{col}5", h, BOLD, fill=FILL_SUB, align="wrap", border=BOX)
wsn.row_dimensions[5].height = 32

n0 = 6
NMONTHS = 60
for m in range(1, NMONTHS + 1):
    rw = n0 + m - 1
    prev = rw - 1
    cumA = "0" if m == 1 else f"E{prev}"
    cumB = "0" if m == 1 else f"H{prev}"
    w(wsn, f"A{rw}", m, BLACK, NUM, align="center")
    w(wsn, f"B{rw}", f"=EDATE({A['start']},A{rw}-1)", BLACK, DATE, align="center")
    w(wsn, f"C{rw}", f"=ROUNDUP(A{rw}/12,0)", BLACK, NUM, align="center")
    # Track A
    w(wsn, f"D{rw}", f"=IF(A{rw}<{A['a_start']},0,MAX(0,MIN({A['a_pace']},{A['a_units']}-{cumA})))",
      BLACK, NUM, align="center")
    w(wsn, f"E{rw}", f"={cumA}+D{rw}", BLACK, NUM, align="center")
    lagA = m - 1 - int(0)   # resolved below via INDEX-free direct row reference
    w(wsn, f"F{rw}", f"=IF(A{rw}-{A['a_lag']}<1,0,INDEX($E${n0}:$E${n0+NMONTHS-1},A{rw}-{A['a_lag']}))",
      BLACK, NUM, align="center")
    # Track B
    w(wsn, f"G{rw}", f"=IF(A{rw}<{A['b_start']},0,MAX(0,MIN({A['b_pace']},{A['b_units']}-{cumB})))",
      BLACK, NUM, align="center")
    w(wsn, f"H{rw}", f"={cumB}+G{rw}", BLACK, NUM, align="center")
    w(wsn, f"I{rw}", f"=IF(A{rw}-{A['b_lag']}<1,0,INDEX($H${n0}:$H${n0+NMONTHS-1},A{rw}-{A['b_lag']}))",
      BLACK, NUM, align="center")
    # combined
    w(wsn, f"J{rw}", f"=E{rw}+H{rw}", BLACK, NUM, align="center")
    w(wsn, f"K{rw}", f"=F{rw}+I{rw}", BLACK, NUM, align="center")
    w(wsn, f"L{rw}", f"=F{rw}*{A['a_prem']}+I{rw}*{A['b_prem']}", BLACK, CUR)
    w(wsn, f"M{rw}", f"=D{rw}*{A['a_cost']}", BLACK, CUR)
    w(wsn, f"N{rw}", f"=G{rw}*{A['b_cost']}", BLACK, CUR)
    w(wsn, f"O{rw}", f"=M{rw}+N{rw}", BLACK, CUR)
    w(wsn, f"P{rw}", f"=(D{rw}*{A['a_down']}+G{rw}*{A['b_down']})/30/{A['units']}", BLACK, PCT2)
nN = n0 + NMONTHS - 1
tr = nN + 1
w(wsn, f"A{tr}", "Total", BOLD, border=TOP)
for col in ["D", "G"]:
    w(wsn, f"{col}{tr}", f"=SUM({col}{n0}:{col}{nN})", BOLD, NUM, border=TOP)
for col in ["M", "N", "O"]:
    w(wsn, f"{col}{tr}", f"=SUM({col}{n0}:{col}{nN})", BOLD, CUR, border=TOP)

sb = tr + 2
w(wsn, f"A{sb}", "ANNUAL SUMMARY — FEEDS THE CASH FLOW MODEL", BOLD, fill=FILL_SUB)
for i, c in enumerate(YR_COLS):
    w(wsn, f"{c}{sb}", f"Year {i+1}", BOLD, fill=FILL_SUB, align="center")
NREF = {}
summary = [
    ("Track A Units Completed in Year", f"=SUMIFS($D${n0}:$D${nN},$C${n0}:$C${nN},{{y}})", NUM, "doneA"),
    ("Track B Units Completed in Year", f"=SUMIFS($G${n0}:$G${nN},$C${n0}:$C${nN},{{y}})", NUM, "doneB"),
    ("Cumulative Renovated at Year End",
     f"=SUMIFS($D${n0}:$D${nN},$C${n0}:$C${nN},\"<=\"&{{y}})+SUMIFS($G${n0}:$G${nN},$C${n0}:$C${nN},\"<=\"&{{y}})",
     NUM, "cum"),
    ("Average Units Earning the Premium",
     f"=AVERAGEIFS($K${n0}:$K${nN},$C${n0}:$C${nN},{{y}})", NUM1, "avg_reno"),
    ("Average Premium in Force ($/month)",
     f"=AVERAGEIFS($L${n0}:$L${nN},$C${n0}:$C${nN},{{y}})", CUR, "prem_force"),
    ("Renovation Capital Spend", f"=SUMIFS($O${n0}:$O${nN},$C${n0}:$C${nN},{{y}})", CUR, "capex"),
    ("Renovation Downtime Vacancy (% of GPR)",
     f"=AVERAGEIFS($P${n0}:$P${nN},$C${n0}:$C${nN},{{y}})", PCT2, "dvac"),
]
for i, (lbl, f, fmt, key) in enumerate(summary):
    rw = sb + 1 + i
    w(wsn, f"A{rw}", lbl, BLACK)
    for j, c in enumerate(YR_COLS):
        w(wsn, f"{c}{rw}", f.replace("{y}", str(j + 1)), BLACK, fmt, align="center")
    NREF[key] = rw
dm = sb + len(summary) + 2
w(wsn, f"A{dm}", "Programme Completion (month all 118 units are renovated)", BOLD)
w(wsn, f"C{dm}", f"=MATCH({A['units']},$J${n0}:$J${nN},0)", BOLD, NUM, align="center", fill=FILL_KEY)
NREF["done_month"] = dm
wsn.merge_cells(f"A{dm+2}:P{dm+2}")
w(wsn, f"A{dm+2}",
  "Track A (53 units with existing laundry) is renovated in place at 10 a month from month 2 — no permit, no "
  "vacancy loss — but the premium only lands at the next renewal, modelled as a 6-month lag. Track B (65 units "
  "gaining laundry) needs a state Construction Design Release and a local DBNS permit and must be vacant, so it "
  "runs at natural turnover of 5 a month from month 4 with 14 days of incremental downtime, and captures the "
  "premium immediately. Splitting the programme this way completes all 118 units far sooner than a single-track "
  "plan without forcing a single resident out.", NOTE, align="wrap")
wsn.row_dimensions[dm+2].height = 56

# ============================================================================
# 9. ANNUAL MODEL
# ============================================================================
wsy = wb.create_sheet("Annual Model")
wsy.sheet_view.showGridLines = False
wsy.column_dimensions["A"].width = 44
wsy.column_dimensions["B"].width = 30
for c in YR_COLS + [EXIT_COL]:
    wsy.column_dimensions[c].width = 14
title_block(wsy, "Five-Year Operating Cash Flow — August 2026 through July 2031")

MR = {}     # model row keys


def m_hdr(row):
    w(wsy, f"A{row}", "", BOLD)
    w(wsy, f"B{row}", "Basis", BOLD, fill=FILL_SUB)
    for i, c in enumerate(YR_COLS):
        w(wsy, f"{c}{row}", f"Year {i+1}", BOLD, fill=FILL_SUB, align="center")
    w(wsy, f"{EXIT_COL}{row}", "Year 6\n(exit)", BOLD, fill=FILL_SUB, align="wrap")
    row += 1
    w(wsy, f"A{row}", "Period Ending", ITAL)
    for i, c in enumerate(YR_COLS):
        w(wsy, f"{c}{row}", f"=EDATE({A['start']},{(i+1)*12}-1)", ITAL, DATE, align="center")
    w(wsy, f"{EXIT_COL}{row}", f"=EDATE({A['start']},72-1)", ITAL, DATE, align="center")
    return row + 1


def m_row(row, label, basis, formulas, fmt, font=BLACK, bold=False, border=None,
          key=None, fill=None, exit_formula=None):
    w(wsy, f"A{row}", label, BOLD if bold else BLACK, border=border, fill=fill)
    w(wsy, f"B{row}", basis, NOTE, align="wrap", border=border)
    for i, c in enumerate(YR_COLS):
        f = formulas[i] if isinstance(formulas, list) else formulas.replace("{c}", c).replace("{y}", str(i + 1))
        w(wsy, f"{c}{row}", f, BOLD if bold else font, fmt, border=border, fill=fill)
    if exit_formula is not None:
        w(wsy, f"{EXIT_COL}{row}", exit_formula, BOLD if bold else font, fmt, border=border, fill=fill)
    if key:
        MR[key] = row
    return row + 1


r = 5
r = m_hdr(r)
r = section(wsy, r, "MARKET RENT INDEXATION")
r = m_row(r, "Market Rent Growth", "Assumptions",
          "=Assumptions!{c}$" + str(AROW["g_rent"]), PCT, font=GREEN, key="g",
          exit_formula=f"=Assumptions!G${AROW['g_rent']}")
_idx_row = r
r = m_row(r, "Rent Index — End of Year",
          "Prior year index × (1 + growth)",
          [f"=1*(1+C{MR['g']})"] +
          [f"={YR_COLS[i-1]}{_idx_row}*(1+{c}{MR['g']})" for i, c in enumerate(YR_COLS[1:], start=1)],
          "0.0000", key="idx", exit_formula=f"=G{_idx_row}*(1+{EXIT_COL}{MR['g']})")
r = m_row(r, "Rent Factor — Mid-Year Convention",
          "SQRT(prior year-end index × this year-end index)",
          [f"=SQRT(1*C{MR['idx']})"] +
          [f"=SQRT({YR_COLS[i-1]}{MR['idx']}*{c}{MR['idx']})" for i, c in enumerate(YR_COLS[1:], start=1)],
          "0.0000", key="fac", exit_formula=f"=SQRT(G{MR['idx']}*{EXIT_COL}{MR['idx']})")

r += 1
r = section(wsy, r, "UNIT POSITION")
RS = "'Reno Schedule'"
r = m_row(r, "Track A Units Completed (in-place)", "Reno Schedule",
          f"={RS}!{{c}}${NREF['doneA']}", NUM, font=GREEN, key="udA", exit_formula="=0")
r = m_row(r, "Track B Units Completed (vacant + laundry)", "Reno Schedule",
          f"={RS}!{{c}}${NREF['doneB']}", NUM, font=GREEN, key="udB", exit_formula="=0")
r = m_row(r, "Cumulative Renovated (% of portfolio)", "Reno Schedule",
          f"={RS}!{{c}}${NREF['cum']}/{A['units']}", PCT, key="upct", exit_formula="=1")
r = m_row(r, "Average Units Earning the Premium", "Reno Schedule",
          f"={RS}!{{c}}${NREF['avg_reno']}", NUM1, font=GREEN, key="ur",
          exit_formula=f"={A['units']}")
r = m_row(r, "Average Premium in Force ($/month)", "Reno Schedule",
          f"={RS}!{{c}}${NREF['prem_force']}", CUR, font=GREEN, key="premf",
          exit_formula=f"={A['units']}*{A['prem_blend']}")

r += 1
r = section(wsy, r, "REVENUE")
r = m_row(r, "Gross Potential Rent — In-Place Market Rents",
          "All 118 units at the 07/01/2026 market rent, indexed for growth",
          f"={A['units']}*{MIXREF['mkt_avg']}*12*{{c}}{MR['fac']}", CUR, key="gpr_c",
          exit_formula=f"={A['units']}*{MIXREF['mkt_avg']}*12*{EXIT_COL}{MR['fac']}")
r = m_row(r, "Gross Potential Rent — Renovation Premium",
          "Premium in force × 12 × rent factor (Track A lags 6 months, Track B 1 month)",
          f"={{c}}{MR['premf']}*12*{{c}}{MR['fac']}", CUR, key="gpr_r",
          exit_formula=f"={EXIT_COL}{MR['premf']}*12*{EXIT_COL}{MR['fac']}")
r = m_row(r, "Gross Potential Rent", "", f"={{c}}{MR['gpr_c']}+{{c}}{MR['gpr_r']}", CUR,
          bold=True, border=TOP, key="gpr",
          exit_formula=f"={EXIT_COL}{MR['gpr_c']}+{EXIT_COL}{MR['gpr_r']}")
r = m_row(r, "   Average Monthly Rent per Unit", "GPR / units / 12",
          f"={{c}}{MR['gpr']}/{A['units']}/12", CUR, font=ITAL, key="avgrent",
          exit_formula=f"={EXIT_COL}{MR['gpr']}/{A['units']}/12")
r = m_row(r, "Less: Loss to Lease", "% of GPR (Assumptions)",
          f"=-{{c}}{MR['gpr']}*Assumptions!{{c}}${AROW['ll']}", CUR, key="ll",
          exit_formula=f"=-{EXIT_COL}{MR['gpr']}*Assumptions!G${AROW['ll']}")
r = m_row(r, "Less: Operating Physical Vacancy", "% of GPR (Assumptions)",
          f"=-{{c}}{MR['gpr']}*Assumptions!{{c}}${AROW['vac']}", CUR, key="vac",
          exit_formula=f"=-{EXIT_COL}{MR['gpr']}*Assumptions!G${AROW['vac']}")
r = m_row(r, "Less: Renovation Downtime Vacancy", "Reno Schedule × GPR",
          f"=-{{c}}{MR['gpr']}*{RS}!{{c}}${NREF['dvac']}", CUR, key="dvac", exit_formula="=0")
r = m_row(r, "Less: Concessions", "% of GPR (Assumptions)",
          f"=-{{c}}{MR['gpr']}*Assumptions!{{c}}${AROW['conc']}", CUR, key="conc",
          exit_formula=f"=-{EXIT_COL}{MR['gpr']}*Assumptions!G${AROW['conc']}")
r = m_row(r, "Less: Bad Debt", "% of GPR (Assumptions)",
          f"=-{{c}}{MR['gpr']}*Assumptions!{{c}}${AROW['bd']}", CUR, key="bd",
          exit_formula=f"=-{EXIT_COL}{MR['gpr']}*Assumptions!G${AROW['bd']}")
r = m_row(r, "Net Rental Income", "",
          f"=SUM({{c}}{MR['gpr']}:{{c}}{MR['bd']})-{{c}}{MR['avgrent']}", CUR,
          bold=True, border=TOP, key="nri",
          exit_formula=f"=SUM({EXIT_COL}{MR['gpr']}:{EXIT_COL}{MR['bd']})-{EXIT_COL}{MR['avgrent']}")
r = m_row(r, "   Economic Occupancy", "NRI / GPR",
          f"={{c}}{MR['nri']}/{{c}}{MR['gpr']}", PCT, font=ITAL, key="econocc",
          exit_formula=f"={EXIT_COL}{MR['nri']}/{EXIT_COL}{MR['gpr']}")
r = m_row(r, "   Physical Occupancy", "1 - operating - downtime vacancy",
          f"=1+({{c}}{MR['vac']}+{{c}}{MR['dvac']})/{{c}}{MR['gpr']}", PCT, font=ITAL,
          exit_formula=f"=1+({EXIT_COL}{MR['vac']}+{EXIT_COL}{MR['dvac']})/{EXIT_COL}{MR['gpr']}")
r = m_row(r, "Other Income", "Other Income tab",
          f"='Other Income'!{{c}}${OI_TOT_ROW}", CUR, font=GREEN, key="oi",
          exit_formula=f"='Other Income'!G${OI_TOT_ROW}*(1+Assumptions!G${AROW['g_oi']})")
r = m_row(r, "Effective Gross Revenue", "", f"={{c}}{MR['nri']}+{{c}}{MR['oi']}", CUR,
          bold=True, border=TOPBOT, key="egr", fill=FILL_TOT,
          exit_formula=f"={EXIT_COL}{MR['nri']}+{EXIT_COL}{MR['oi']}")

r += 1
r = section(wsy, r, "OPERATING EXPENSES")


def exp_line(row, label, key_a, mrkey, basis):
    formulas = [f"={A[key_a]}"]
    for i, c in enumerate(YR_COLS[1:], start=1):
        formulas.append(f"={YR_COLS[i-1]}{row}*(1+{A[key_a + '_g']})")
    return m_row(row, label, basis, formulas, CUR, key=mrkey,
                 exit_formula=f"=G{row}*(1+{A[key_a + '_g']})")


r = exp_line(r, "Payroll", "e_pay", "x_pay", "Year 1 input, grown")
r = m_row(r, "Property Management Fee", "% of effective gross revenue",
          f"={{c}}{MR['egr']}*{A['e_mgmt']}", CUR, key="mgmt",
          exit_formula=f"={EXIT_COL}{MR['egr']}*{A['e_mgmt']}")
r = exp_line(r, "Marketing & Advertising", "e_mkt", "x_mkt", "Year 1 input, grown")
r = exp_line(r, "General & Administrative", "e_ga", "x_ga", "Year 1 input, grown")
r = exp_line(r, "Bulk Internet", "e_int", "x_int", "$27.95/unit/month stabilised, grown")
r = exp_line(r, "Electricity", "e_elec", "x_elec", "Year 1 input, grown")
r = exp_line(r, "Gas", "e_gas", "x_gas", "Year 1 input, grown")
r = exp_line(r, "Water & Sewer", "e_ws", "x_ws", "Year 1 input, grown; 85% recovered in Other Income")
r = exp_line(r, "Repairs & Maintenance", "e_rm", "x_rm", "Year 1 input, grown")
r = exp_line(r, "Insurance", "e_ins", "x_ins", "Year 1 input, grown")
r = m_row(r, "Real Estate Taxes", "Explicit annual schedule (reassessment on sale)",
          f"=Assumptions!{{c}}${AROW['e_tax']}", CUR, font=GREEN, key="tax",
          exit_formula=f"=Assumptions!G${AROW['e_tax']}*1.05")
r = m_row(r, "Total Operating Expenses", "",
          f"=SUM({{c}}{MR['x_pay']}:{{c}}{MR['tax']})", CUR, bold=True, border=TOP, key="opex",
          exit_formula=f"=SUM({EXIT_COL}{MR['x_pay']}:{EXIT_COL}{MR['tax']})")
r = m_row(r, "   Per Unit", "", f"={{c}}{MR['opex']}/{A['units']}", CUR, font=ITAL,
          exit_formula=f"={EXIT_COL}{MR['opex']}/{A['units']}")
r = m_row(r, "   Operating Expense Ratio", "Total opex / EGR",
          f"={{c}}{MR['opex']}/{{c}}{MR['egr']}", PCT, font=ITAL,
          exit_formula=f"={EXIT_COL}{MR['opex']}/{EXIT_COL}{MR['egr']}")

r += 1
r = m_row(r, "NET OPERATING INCOME", "", f"={{c}}{MR['egr']}-{{c}}{MR['opex']}", CUR,
          bold=True, border=TOPBOT, key="noi", fill=FILL_KEY,
          exit_formula=f"={EXIT_COL}{MR['egr']}-{EXIT_COL}{MR['opex']}")
r = m_row(r, "   NOI per Unit", "", f"={{c}}{MR['noi']}/{A['units']}", CUR, font=ITAL,
          exit_formula=f"={EXIT_COL}{MR['noi']}/{A['units']}")
r = m_row(r, "   NOI Margin", "", f"={{c}}{MR['noi']}/{{c}}{MR['egr']}", PCT, font=ITAL)
r = m_row(r, "   NOI Growth", "",
          ["=\"n/a\""] + [f"={c}{MR['noi']}/{YR_COLS[i-1]}{MR['noi']}-1" for i, c in enumerate(YR_COLS[1:], start=1)],
          PCT, font=ITAL)
r = m_row(r, "   Yield on Purchase Price", "NOI / purchase price",
          f"={{c}}{MR['noi']}/{A['price']}", PCT, font=ITAL, key="yoc")
r = m_row(r, "   Yield on Total Cost", "NOI / (price + all capital)",
          f"={{c}}{MR['noi']}/Returns!$C${{tcost}}", PCT, font=ITAL, key="yotc")

r += 1
r = section(wsy, r, "BELOW NOI")
r = m_row(r, "Less: Replacement Reserves", "$/unit/year, grown",
          [f"=-{A['e_res']}*{A['units']}"] +
          [f"={YR_COLS[i-1]}{r}*(1+Assumptions!{c}${AROW['g_oi']})"
           for i, c in enumerate(YR_COLS[1:], start=1)], CUR, key="res")
r = m_row(r, "Net Cash Flow before Debt Service", "",
          f"={{c}}{MR['noi']}+{{c}}{MR['res']}", CUR, bold=True, border=TOP, key="cfbd")
r = m_row(r, "Less: Interest Expense", "Loan balance × rate (interest-only)",
          f"=-Returns!$C${{debt}}*{A['rate']}", CUR, key="int")
r = m_row(r, "Cash Flow to Equity (operations)", "",
          f"={{c}}{MR['cfbd']}+{{c}}{MR['int']}", CUR, bold=True, border=TOPBOT, key="cfe", fill=FILL_TOT)
r = m_row(r, "   Debt Service Coverage Ratio", "NOI / interest",
          f"={{c}}{MR['noi']}/-{{c}}{MR['int']}", MULT, font=ITAL)
r = m_row(r, "   Debt Yield", "NOI / loan balance",
          f"={{c}}{MR['noi']}/Returns!$C${{debt}}", PCT, font=ITAL)
r = m_row(r, "   Cash-on-Cash Return", "CF to equity / sponsor equity",
          f"={{c}}{MR['cfe']}/Returns!$C${{eq}}", PCT, font=ITAL, key="coc")

r += 1
r = section(wsy, r, "MEMO — CAPITAL SPEND (funded at closing per Sources & Uses)")
r = m_row(r, "Renovation Capital", "Reno Schedule", f"={RS}!{{c}}${NREF['capex']}", CUR,
          font=GREEN, key="cx_reno")
r = m_row(r, "Non-Unit / Deferred Maintenance Capital", "Capex Budget",
          ["='Capex Budget'!D" + str(CAP_TOT_ROW), "='Capex Budget'!E" + str(CAP_TOT_ROW),
           "='Capex Budget'!F" + str(CAP_TOT_ROW), "='Capex Budget'!G" + str(CAP_TOT_ROW),
           "='Capex Budget'!H" + str(CAP_TOT_ROW)], CUR, font=GREEN, key="cx_prop")
r = m_row(r, "Contingency", "% of capital spend",
          f"=({{c}}{MR['cx_reno']}+{{c}}{MR['cx_prop']})*{A['conting']}", CUR, key="cx_cont")
r = m_row(r, "Total Capital Deployed", "",
          f"=SUM({{c}}{MR['cx_reno']}:{{c}}{MR['cx_cont']})", CUR, bold=True, border=TOP, key="cx_tot")
CX_TOT = MR["cx_tot"]
r += 1
w(wsy, f"A{r}",
  "Capital is deployed over the schedule shown above but funded in full at closing in Sources & Uses, so it is "
  "not deducted again from annual cash flow. That convention is conservative: it puts 100% of the equity out on "
  "day one and gives the model no benefit from the real timing of draws.", NOTE, align="wrap")
wsy.row_dimensions[r].height = 34

# patch T-12 Recast cross references now that Annual Model rows are known
for ph, row in [("{oitot}", OI_TOT_ROW), ("{gpr}", MR["gpr"]), ("{ll}", MR["ll"]), ("{vac}", MR["vac"]), ("{dvac}", MR["dvac"]),
                ("{conc}", MR["conc"]), ("{bd}", MR["bd"]), ("{nri}", MR["nri"]),
                ("{mgmt}", MR["mgmt"]), ("{tax}", MR["tax"]), ("{opex}", MR["opex"]),
                ("{noi}", MR["noi"])]:
    for rr in range(1, wst.max_row + 1):
        for cc in range(1, 7):
            cell = wst.cell(row=rr, column=cc)
            if isinstance(cell.value, str) and ph in cell.value:
                cell.value = cell.value.replace(ph, str(row))

# patch Other Income % of EGR
for c in YR_COLS:
    wso[f"{c}{OI_PCT_ROW}"] = f"={c}{OI_TOT_ROW}/'Annual Model'!{c}{MR['egr']}"

wsa[f"B{AROW['done_month']}"] = f"='Reno Schedule'!$C${NREF['done_month']}"
# patch the Assumptions going-in cap now that the Annual Model NOI row is known
wsa[f"B{AROW['goingin']}"] = f"='Annual Model'!C{MR['noi']}/{A['price']}"

# ============================================================================
# 10. RETURNS
# ============================================================================
wsq = wb.create_sheet("Returns")
wsq.sheet_view.showGridLines = False
wsq.column_dimensions["A"].width = 44
wsq.column_dimensions["B"].width = 13
wsq.column_dimensions["C"].width = 15
for c in ["D", "E", "F", "G", "H", "I"]:
    wsq.column_dimensions[c].width = 14
wsq.column_dimensions["J"].width = 56
title_block(wsq, "Sources & Uses, Debt, Exit Valuation and Returns", "J")

QR = {}
r = 5
r = section(wsq, r, "SOURCES & USES", "J")
w(wsq, f"A{r}", "USES", BOLD, fill=FILL_SUB)
w(wsq, f"C{r}", "Amount", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"D{r}", "Per Unit", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"E{r}", "% of Total", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"J{r}", "Note", BOLD, fill=FILL_SUB)
r += 1
u0 = r
uses = [
    ("Purchase Price", f"={A['price']}", "Implied by the OM's 6.9% Year-1 cap rate on CBRE's forecast NOI."),
    ("Acquisition Costs", f"={A['price']}*{A['acqcost']}", "1.5% of price."),
    ("Renovation Capital (118 units)", f"={MIXREF['reno_cap']}",
     "$13,000/unit interior plus $5,000/unit for the 65 units gaining laundry."),
    ("Non-Unit / Deferred Maintenance Capital", f"='Capex Budget'!B{CAP_TOT_ROW}",
     "Roofs, asphalt, sump pumps, glass, entries, amenities."),
    ("Capital Contingency", None, "8% of all capital."),
    ("Financing Costs", None, "1% origination plus lender legal, appraisal, engineering and escrow set-up."),
]
for lbl, f, note in uses:
    w(wsq, f"A{r}", lbl, BLACK)
    w(wsq, f"C{r}", f, BLACK if f else BLACK, CUR)
    w(wsq, f"D{r}", f"=C{r}/{A['units']}", BLACK, CUR)
    w(wsq, f"J{r}", note, NOTE, align="wrap")
    r += 1
uN = r - 1
QR["cont_row"] = uN - 1
QR["fin_row"] = uN
w(wsq, f"A{r}", "Total Uses", BOLD, border=TOP)
w(wsq, f"C{r}", f"=SUM(C{u0}:C{uN})", BOLD, CUR, border=TOP)
w(wsq, f"D{r}", f"=C{r}/{A['units']}", BOLD, CUR, border=TOP)
QR["uses"] = r
for rr in range(u0, uN + 1):
    w(wsq, f"E{rr}", f"=C{rr}/$C${QR['uses']}", BLACK, PCT)
w(wsq, f"E{r}", f"=SUM(E{u0}:E{uN})", BOLD, PCT, border=TOP)
r += 2

w(wsq, f"A{r}", "SOURCES", BOLD, fill=FILL_SUB)
w(wsq, f"C{r}", "Amount", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"D{r}", "Per Unit", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"E{r}", "% of Total", BOLD, fill=FILL_SUB, align="center")
r += 1
s0 = r
w(wsq, f"A{r}", "Senior / Value-Add Loan", BLACK)
w(wsq, f"C{r}", None, BLACK, CUR)
QR["debt"] = r
r += 1
w(wsq, f"A{r}", "Sponsor Equity", BLACK)
w(wsq, f"C{r}", f"=C{QR['uses']}-C{QR['debt']}", BOLD, CUR)
QR["eq"] = r
r += 1
sN = r - 1
w(wsq, f"A{r}", "Total Sources", BOLD, border=TOP)
w(wsq, f"C{r}", f"=SUM(C{s0}:C{sN})", BOLD, CUR, border=TOP)
w(wsq, f"D{r}", f"=C{r}/{A['units']}", BOLD, CUR, border=TOP)
QR["sources"] = r
for rr in [QR["debt"], QR["eq"]]:
    w(wsq, f"D{rr}", f"=C{rr}/{A['units']}", BLACK, CUR)
    w(wsq, f"E{rr}", f"=C{rr}/$C${QR['sources']}", BLACK, PCT)
w(wsq, f"E{r}", f"=SUM(E{s0}:E{sN})", BOLD, PCT, border=TOP)
r += 1
w(wsq, f"A{r}", "   Check: sources = uses", ITAL)
w(wsq, f"C{r}", f"=C{QR['sources']}-C{QR['uses']}", BLACK, CUR)
r += 2

r = section(wsq, r, "TOTAL CAPITALISATION & DEBT TERMS", "J")
tcap = [
    ("Total Capitalised Cost (price + all capital)",
     f"={A['price']}+{MIXREF['reno_cap']}+'Capex Budget'!B{CAP_TOT_ROW}+C{QR['cont_row']}", CUR, "tcost",
     "Basis for loan-to-cost sizing and the yield-on-cost metric."),
    ("   Total Cost per Unit", None, CUR, "tcostpu", ""),
    ("Loan Sizing Test 1 — Maximum Loan to Cost", None, CUR, "t_ltc",
     f"Maximum loan-to-cost applied to total capitalised cost."),
    ("Loan Sizing Test 2 — Minimum Debt Yield", None, CUR, "t_dy",
     "Year-1 NOI divided by the minimum debt yield."),
    ("Loan Sizing Test 3 — Minimum DSCR", None, CUR, "t_dscr",
     "Year-1 NOI divided by (minimum DSCR × interest rate), interest-only."),
    ("Loan Amount (lesser of the three tests)", None, CUR, "loan",
     "Sizing to the lesser of the three tests, as a lender would."),
    ("   Effective Loan to Cost", None, PCT, None, ""),
    ("   Loan per Unit", None, CUR, None, ""),
    ("Interest Rate (fixed, interest-only)", f"={A['rate']}", PCT2, None, ""),
    ("Annual Interest Expense", None, CUR, "annint", ""),
    ("Year-1 Debt Yield", None, PCT, None, "Lender test: NOI / loan."),
    ("Year-1 DSCR", None, MULT, None, "Lender test: NOI / interest."),
]
for lbl, f, fmt, key, note in tcap:
    w(wsq, f"A{r}", lbl, BLACK)
    if f:
        w(wsq, f"C{r}", f, BLACK, fmt)
    if key:
        QR[key] = r
    if note:
        w(wsq, f"J{r}", note, NOTE, align="wrap")
    r += 1
tc = QR["tcost"]
w(wsq, f"C{tc+1}", f"=C{tc}/{A['units']}", BLACK, CUR)
w(wsq, f"C{QR['t_ltc']}", f"=C{tc}*{A['ltc']}", BLACK, CUR)
w(wsq, f"C{QR['t_dy']}", f"='Annual Model'!C{MR['noi']}/{A['mindy']}", BLACK, CUR)
w(wsq, f"C{QR['t_dscr']}", f"='Annual Model'!C{MR['noi']}/({A['mindscr']}*{A['rate']})", BLACK, CUR)
w(wsq, f"C{QR['loan']}", f"=MIN(C{QR['t_ltc']},C{QR['t_dy']},C{QR['t_dscr']})", BOLD, CUR, fill=FILL_TOT)
w(wsq, f"C{QR['loan']+1}", f"=C{QR['loan']}/C{tc}", BLACK, PCT)
w(wsq, f"C{QR['loan']+2}", f"=C{QR['loan']}/{A['units']}", BLACK, CUR)
w(wsq, f"C{QR['annint']}", f"=C{QR['loan']}*{A['rate']}", BLACK, CUR)
w(wsq, f"C{QR['annint']+1}", f"='Annual Model'!C{MR['noi']}/C{QR['loan']}", BLACK, PCT)
w(wsq, f"C{QR['annint']+2}", f"='Annual Model'!C{MR['noi']}/C{QR['annint']}", BLACK, MULT)
# now fill the deferred Sources & Uses cells
wsq[f"C{QR['cont_row']}"] = (f"=({MIXREF['reno_cap']}+'Capex Budget'!B{CAP_TOT_ROW})*{A['conting']}")
wsq[f"C{QR['cont_row']}"].font = BLACK
wsq[f"C{QR['cont_row']}"].number_format = CUR
wsq[f"C{QR['fin_row']}"] = f"=C{QR['loan']}*{A['fee']}+{A['fincost']}"
wsq[f"C{QR['fin_row']}"].font = BLACK
wsq[f"C{QR['fin_row']}"].number_format = CUR
wsq[f"C{QR['debt']}"] = f"=C{QR['loan']}"
wsq[f"C{QR['debt']}"].font = BOLD
wsq[f"C{QR['debt']}"].number_format = CUR

r += 1
r = section(wsq, r, "EXIT VALUATION — END OF YEAR 5", "J")
ex = [
    ("Year-6 Forward NOI", f"='Annual Model'!{EXIT_COL}{MR['noi']}", CUR, "fwdnoi",
     "Next-twelve-months NOI a buyer would capitalise."),
    ("Exit Cap Rate", f"={A['exitcap']}", PCT2, None,
     "15 bps inside the going-in cap, reflecting a fully renovated asset with in-unit laundry throughout."),
    ("Gross Sale Value", None, CUR, "gross", ""),
    ("   Value per Unit", None, CUR, None, ""),
    ("Less: Cost of Sale", None, CUR, "cos", ""),
    ("Net Sale Proceeds", None, CUR, "net", ""),
    ("Less: Loan Repayment", None, CUR, "repay", ""),
    ("Net Proceeds to Equity", None, CUR, "toeq", ""),
    ("   Value Created over Total Cost", None, CUR, None,
     "Net sale proceeds less total capitalised cost and acquisition costs."),
]
for lbl, f, fmt, key, note in ex:
    w(wsq, f"A{r}", lbl, BLACK)
    if f:
        w(wsq, f"C{r}", f, BLACK, fmt)
    if key:
        QR[key] = r
    if note:
        w(wsq, f"J{r}", note, NOTE, align="wrap")
    r += 1
w(wsq, f"C{QR['gross']}", f"=C{QR['fwdnoi']}/{A['exitcap']}", BOLD, CUR)
w(wsq, f"C{QR['gross']+1}", f"=C{QR['gross']}/{A['units']}", BLACK, CUR)
w(wsq, f"C{QR['cos']}", f"=-C{QR['gross']}*{A['cos']}", BLACK, CUR)
w(wsq, f"C{QR['net']}", f"=C{QR['gross']}+C{QR['cos']}", BOLD, CUR)
w(wsq, f"C{QR['repay']}", f"=-C{QR['loan']}", BLACK, CUR)
w(wsq, f"C{QR['toeq']}", f"=C{QR['net']}+C{QR['repay']}", BOLD, CUR)
w(wsq, f"C{QR['toeq']+1}", f"=C{QR['net']}-C{tc}-{A['price']}*{A['acqcost']}", BLACK, CUR)

r += 1
r = section(wsq, r, "CASH FLOWS AND RETURNS", "J")
w(wsq, f"A{r}", "", BOLD)
w(wsq, f"C{r}", "At Close", BOLD, fill=FILL_SUB, align="center")
for i, c in enumerate(["D", "E", "F", "G", "H"]):
    w(wsq, f"{c}{r}", f"Year {i+1}", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"I{r}", "Total", BOLD, fill=FILL_SUB, align="center")
r += 1
CFCOLS = ["D", "E", "F", "G", "H"]

w(wsq, f"A{r}", "UNLEVERED", BOLD)
r += 1
w(wsq, f"A{r}", "Net Cash Flow before Debt Service", BLACK)
w(wsq, f"C{r}", f"=-(C{tc}+{A['price']}*{A['acqcost']})", BLACK, CUR)
for i, c in enumerate(CFCOLS):
    w(wsq, f"{c}{r}", f"='Annual Model'!{YR_COLS[i]}{MR['cfbd']}", GREEN, CUR)
QR["ucf_op"] = r
r += 1
w(wsq, f"A{r}", "Net Sale Proceeds", BLACK)
w(wsq, f"H{r}", f"=C{QR['net']}", BLACK, CUR)
QR["ucf_sale"] = r
r += 1
w(wsq, f"A{r}", "Total Unlevered Cash Flow", BOLD, border=TOP)
w(wsq, f"C{r}", f"=C{QR['ucf_op']}", BOLD, CUR, border=TOP)
for c in CFCOLS:
    w(wsq, f"{c}{r}", f"=SUM({c}{QR['ucf_op']}:{c}{QR['ucf_sale']})", BOLD, CUR, border=TOP)
w(wsq, f"I{r}", f"=SUM(C{r}:H{r})", BOLD, CUR, border=TOP)
QR["ucf"] = r
r += 2

w(wsq, f"A{r}", "LEVERED", BOLD)
r += 1
w(wsq, f"A{r}", "Sponsor Equity", BLACK)
w(wsq, f"C{r}", f"=-C{QR['eq']}", BLACK, CUR)
QR["lcf_eq"] = r
r += 1
w(wsq, f"A{r}", "Cash Flow to Equity (operations)", BLACK)
for i, c in enumerate(CFCOLS):
    w(wsq, f"{c}{r}", f"='Annual Model'!{YR_COLS[i]}{MR['cfe']}", GREEN, CUR)
QR["lcf_op"] = r
r += 1
w(wsq, f"A{r}", "Net Proceeds to Equity at Sale", BLACK)
w(wsq, f"H{r}", f"=C{QR['toeq']}", BLACK, CUR)
QR["lcf_sale"] = r
r += 1
w(wsq, f"A{r}", "Total Levered Cash Flow", BOLD, border=TOP)
w(wsq, f"C{r}", f"=C{QR['lcf_eq']}", BOLD, CUR, border=TOP)
for c in CFCOLS:
    w(wsq, f"{c}{r}", f"=SUM({c}{QR['lcf_op']}:{c}{QR['lcf_sale']})", BOLD, CUR, border=TOP)
w(wsq, f"I{r}", f"=SUM(C{r}:H{r})", BOLD, CUR, border=TOP)
QR["lcf"] = r
r += 1
w(wsq, f"A{r}", "   Cash-on-Cash Return", ITAL)
for i, c in enumerate(CFCOLS):
    w(wsq, f"{c}{r}", f"={c}{QR['lcf_op']}/C{QR['eq']}", ITAL, PCT)
r += 2

r = section(wsq, r, "RETURN SUMMARY", "J")
w(wsq, f"A{r}", "", BOLD)
w(wsq, f"C{r}", "Unlevered", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"D{r}", "Levered", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"J{r}", "Note", BOLD, fill=FILL_SUB)
r += 1
w(wsq, f"A{r}", "Internal Rate of Return", BOLD)
w(wsq, f"C{r}", f"=IRR(C{QR['ucf']}:H{QR['ucf']})", BOLD, PCT, fill=FILL_KEY)
w(wsq, f"D{r}", f"=IRR(C{QR['lcf']}:H{QR['lcf']})", BOLD, PCT, fill=FILL_KEY)
QR["irr"] = r
r += 1
w(wsq, f"A{r}", "Equity Multiple", BOLD)
w(wsq, f"C{r}", f"=SUM(D{QR['ucf']}:H{QR['ucf']})/-C{QR['ucf']}", BOLD, MULT)
w(wsq, f"D{r}", f"=SUM(D{QR['lcf']}:H{QR['lcf']})/-C{QR['lcf']}", BOLD, MULT)
QR["em"] = r
r += 1
w(wsq, f"A{r}", "Profit", BLACK)
w(wsq, f"C{r}", f"=I{QR['ucf']}", BLACK, CUR)
w(wsq, f"D{r}", f"=I{QR['lcf']}", BLACK, CUR)
r += 1
w(wsq, f"A{r}", "Average Cash-on-Cash (Years 1-5)", BLACK)
w(wsq, f"D{r}", f"=AVERAGE(D{QR['lcf_op']}:H{QR['lcf_op']})/C{QR['eq']}", BLACK, PCT)
r += 1
w(wsq, f"A{r}", "Going-In Cap Rate (our underwriting)", BLACK)
w(wsq, f"C{r}", f"='Annual Model'!C{MR['noi']}/{A['price']}", BLACK, PCT2)
r += 1
w(wsq, f"A{r}", "Stabilised Yield on Total Cost (Year 5)", BLACK)
w(wsq, f"C{r}", f"='Annual Model'!G{MR['noi']}/C{tc}", BLACK, PCT2)
w(wsq, f"J{r}", "Development spread over the 6.75% exit cap is the value created by the business plan.",
  NOTE, align="wrap")
r += 1
w(wsq, f"A{r}", "Renovation Return on Cost", BLACK)
w(wsq, f"C{r}", f"={A['prem_blend']}*12/{A['cost_blend']}", BLACK, PCT2)
r += 1
r = section(wsq, r, "HURDLE TEST & RECOMMENDATION", "J")
w(wsq, f"A{r}", "", BOLD)
w(wsq, f"C{r}", "Target", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"D{r}", "This Deal", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"E{r}", "Result", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"J{r}", "Note", BOLD, fill=FILL_SUB)
r += 1
w(wsq, f"A{r}", "Levered IRR", BLACK)
w(wsq, f"C{r}", 0.14, BLUE, PCT, align="center")
w(wsq, f"D{r}", f"=D{QR['irr']}", BLACK, PCT, align="center")
w(wsq, f"E{r}", f'=IF(D{r}>=C{r},"PASS","FAIL")', BOLD, align="center")
hr1 = r
r += 1
w(wsq, f"A{r}", "Equity Multiple", BLACK)
w(wsq, f"C{r}", 1.80, BLUE, MULT, align="center")
w(wsq, f"D{r}", f"=D{QR['em']}", BLACK, MULT, align="center")
w(wsq, f"E{r}", f'=IF(D{r}>=C{r},"PASS","FAIL")', BOLD, align="center")
r += 1
w(wsq, f"A{r}", "Year-5 Yield on Total Cost vs Exit Cap (spread)", BLACK)
w(wsq, f"C{r}", 0.0075, BLUE, PCT, align="center")
w(wsq, f"D{r}", f"='Annual Model'!G{MR['noi']}/C{tc}-{A['exitcap']}", BLACK, PCT, align="center")
w(wsq, f"E{r}", f'=IF(D{r}>=C{r},"PASS","FAIL")', BOLD, align="center")
r += 1
w(wsq, f"A{r}", "Recommended Maximum Bid (the price this model solves for)", BOLD, fill=FILL_KEY)
w(wsq, f"C{r}", f"={A['price']}", GREEN, CUR, fill=FILL_KEY)
QR["bid"] = r
w(wsq, f"D{r}", f"=C{r}/{A['units']}", BOLD, CUR, fill=FILL_KEY)
w(wsq, f"J{r}",
  "The highest basis at which this business plan clears a 14% levered IRR and a 1.80x multiple on the "
  "conservative case. Set it on the Assumptions tab and the whole model, including both sensitivity grids "
  "and all three breakeven ladders, follows.", NOTE, align="wrap")
r += 1
w(wsq, f"A{r}", "Implied Going-In Cap at the Recommended Bid", BLACK)
w(wsq, f"C{r}", f"='Annual Model'!C{MR['noi']}/C{r-1}", BLACK, PCT2)
r += 1
w(wsq, f"A{r}", "Actual Trade Price, 12 June 2026", BLACK)
w(wsq, f"C{r}", f"={A['actual']}", GREEN, CUR)
w(wsq, f"J{r}", "Marion County deed, special warranty, assessor validity flag 'Y'. $115,466 per unit.",
  NOTE, align="wrap")
r += 1
w(wsq, f"A{r}", "  Our Bid vs the Actual Trade", BLACK)
w(wsq, f"C{r}", f"={A['price']}/{A['actual']}-1", BLACK, PCT)
r += 1
r = section(wsq, r, "OUR BID vs WHAT IT ACTUALLY TRADED FOR", "J")
w(wsq, f"A{r}", "", BOLD)
w(wsq, f"C{r}", "At Our Bid", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"D{r}", "At Actual Trade", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"J{r}", "Operating NOI is identical in both cases — only the basis, the loan and the equity change.",
  NOTE, align="wrap")
r += 1
BID = f"{A['actual']}"
bid_rows = [
    ("Purchase Price", f"={A['price']}", f"={BID}", CUR),
    ("Price per Unit", f"={A['price']}/{A['units']}", f"={BID}/{A['units']}", CUR),
    ("Total Capitalised Cost", f"=C{tc}",
     f"={BID}+{MIXREF['reno_cap']}+'Capex Budget'!B{CAP_TOT_ROW}+C{QR['cont_row']}", CUR),
    ("All-in Basis per Unit", f"=C{tc}/{A['units']}", None, CUR),
    ("Going-In Cap Rate (our Year-1 NOI)", f"='Annual Model'!C{MR['noi']}/{A['price']}",
     f"='Annual Model'!C{MR['noi']}/{BID}", PCT2),
    ("Year-5 Yield on Total Cost", f"='Annual Model'!G{MR['noi']}/C{tc}", None, PCT2),
    ("Spread to the Exit Cap", None, None, PCT),
    ("Loan Amount", f"=C{QR['loan']}", None, CUR),
    ("Sponsor Equity", f"=C{QR['eq']}", None, CUR),
    ("Annual Interest Expense", f"=C{QR['annint']}", None, CUR),
    ("Net Proceeds to Equity at Sale", f"=C{QR['toeq']}", None, CUR),
]
bb = r
for lbl, fc, fd, fmt in bid_rows:
    w(wsq, f"A{r}", lbl, BLACK)
    if fc:
        w(wsq, f"C{r}", fc, BLACK, fmt)
    if fd:
        w(wsq, f"D{r}", fd, BLACK, fmt)
    r += 1
w(wsq, f"D{bb+3}", f"=D{bb+2}/{A['units']}", BLACK, CUR)
w(wsq, f"D{bb+5}", f"='Annual Model'!G{MR['noi']}/D{bb+2}", BLACK, PCT2)
w(wsq, f"C{bb+6}", f"=C{bb+5}-{A['exitcap']}", BLACK, PCT)
w(wsq, f"D{bb+6}", f"=D{bb+5}-{A['exitcap']}", BLACK, PCT)
w(wsq, f"D{bb+7}", f"=MIN(D{bb+2}*{A['ltc']},C{QR['t_dy']},C{QR['t_dscr']})", BLACK, CUR)
w(wsq, f"D{bb+8}", f"=D{bb}+D{bb}*{A['acqcost']}+{MIXREF['reno_cap']}+'Capex Budget'!B{CAP_TOT_ROW}"
                   f"+C{QR['cont_row']}+D{bb+7}*{A['fee']}+{A['fincost']}-D{bb+7}", BOLD, CUR)
w(wsq, f"D{bb+9}", f"=D{bb+7}*{A['rate']}", BLACK, CUR)
w(wsq, f"D{bb+10}", f"=C{QR['net']}-D{bb+7}", BLACK, CUR)
r += 1
# cash flow rows
w(wsq, f"A{r}", "Unlevered Cash Flow at Our Bid", BLACK)
w(wsq, f"C{r}", f"=-(D{bb+2}+D{bb}*{A['acqcost']})", BLACK, CUR)
for i, c in enumerate(CFCOLS):
    f_ = f"='Annual Model'!{YR_COLS[i]}{MR['cfbd']}"
    if i == 4:
        f_ += f"+C{QR['net']}"
    w(wsq, f"{c}{r}", f_, BLACK, CUR)
bid_u = r
r += 1
w(wsq, f"A{r}", "Levered Cash Flow at Our Bid", BLACK)
w(wsq, f"C{r}", f"=-D{bb+8}", BLACK, CUR)
for i, c in enumerate(CFCOLS):
    f_ = f"='Annual Model'!{YR_COLS[i]}{MR['cfbd']}-$D${bb+9}"
    if i == 4:
        f_ += f"+D{bb+10}"
    w(wsq, f"{c}{r}", f_, BLACK, CUR)
bid_l = r
r += 1
w(wsq, f"A{r}", "Unlevered IRR  /  Equity Multiple", BOLD, border=TOP)
w(wsq, f"C{r}", f"=IRR(C{QR['ucf']}:H{QR['ucf']})", BOLD, PCT, border=TOP)
w(wsq, f"D{r}", f"=IRR(C{bid_u}:H{bid_u})", BOLD, PCT, border=TOP, fill=FILL_KEY)
w(wsq, f"E{r}", f"=SUM(D{bid_u}:H{bid_u})/-C{bid_u}", BOLD, MULT, border=TOP)
r += 1
w(wsq, f"A{r}", "Levered IRR  /  Equity Multiple", BOLD, border=TOPBOT)
w(wsq, f"C{r}", f"=IRR(C{QR['lcf']}:H{QR['lcf']})", BOLD, PCT, border=TOPBOT)
w(wsq, f"D{r}", f"=IRR(C{bid_l}:H{bid_l})", BOLD, PCT, border=TOPBOT, fill=FILL_KEY)
w(wsq, f"E{r}", f"=SUM(D{bid_l}:H{bid_l})/-C{bid_l}", BOLD, MULT, border=TOPBOT, fill=FILL_KEY)
r += 1
w(wsq, f"A{r}", "Average Cash-on-Cash (Years 1-5)", BLACK)
w(wsq, f"C{r}", f"=AVERAGE(D{QR['lcf_op']}:H{QR['lcf_op']})/C{QR['eq']}", BLACK, PCT)
w(wsq, f"D{r}", f"=(AVERAGE('Annual Model'!C{MR['cfbd']}:G{MR['cfbd']})-D{bb+9})/D{bb+8}", BLACK, PCT)
r += 1
r = section(wsq, r, "VALUE CREATION BRIDGE — WHERE THE MONEY IS MADE (AND LOST)", "J")
vb = r
bridge = [
    ("Purchase Price", f"=-{A['price']}", "What we pay."),
    ("Acquisition Costs", f"=-{A['price']}*{A['acqcost']}", ""),
    ("Renovation Capital (118 units)", f"=-{MIXREF['reno_cap']}", "Earns a 12.7% return on cost."),
    ("Non-Unit / Deferred Maintenance Capital",
     f"=-'Capex Budget'!B{CAP_TOT_ROW}", "Earns no direct rent. Required to hold the asset and to support the "
     "renovated rents, but it is a pure drag on returns."),
    ("Capital Contingency", f"=-C{QR['cont_row']}", ""),
    ("Cumulative NOI, Years 1-5", f"=SUM('Annual Model'!C{MR['noi']}:G{MR['noi']})",
     "Five years of operations."),
    ("Replacement Reserves over the hold", f"=SUM('Annual Model'!C{MR['res']}:G{MR['res']})", ""),
    ("Net Sale Proceeds", f"=C{QR['net']}", "Year-6 NOI capitalised at the exit cap, less 1.5% cost of sale."),
]
w(wsq, f"A{r}", "Component", BOLD, fill=FILL_SUB)
w(wsq, f"C{r}", "Amount", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"D{r}", "Per Unit", BOLD, fill=FILL_SUB, align="center")
w(wsq, f"J{r}", "Comment", BOLD, fill=FILL_SUB)
r += 1
b0 = r
for lbl, f, note in bridge:
    w(wsq, f"A{r}", lbl, BLACK)
    w(wsq, f"C{r}", f, BLACK, CUR)
    w(wsq, f"D{r}", f"=C{r}/{A['units']}", BLACK, CUR)
    if note:
        w(wsq, f"J{r}", note, NOTE, align="wrap")
    r += 1
bN = r - 1
w(wsq, f"A{r}", "Unlevered Profit", BOLD, border=TOPBOT)
w(wsq, f"C{r}", f"=SUM(C{b0}:C{bN})", BOLD, CUR, border=TOPBOT)
w(wsq, f"D{r}", f"=C{r}/{A['units']}", BOLD, CUR, border=TOPBOT)
r += 2
w(wsq, f"A{r}", "NOI lift created by the renovation programme", BLACK)
w(wsq, f"C{r}", f"={A['prem_blend']}*{A['units']}*12*(1-Assumptions!G${AROW['ll']}"
                f"-Assumptions!G${AROW['vac']}-Assumptions!G${AROW['conc']}"
                f"-Assumptions!G${AROW['bd']})*(1-{A['e_mgmt']})", BLACK, CUR)
QR["renoNOI"] = r
w(wsq, f"J{r}", "Blended premium, net of vacancy, loss to lease, concessions, bad debt and the management fee.",
  NOTE, align="wrap")
r += 1
w(wsq, f"A{r}", "Value of that lift at the exit cap", BLACK)
w(wsq, f"C{r}", f"=C{QR['renoNOI']}/{A['exitcap']}", BLACK, CUR)
r += 1
w(wsq, f"A{r}", "Less: renovation capital + its share of contingency", BLACK)
w(wsq, f"C{r}", f"=-{MIXREF['reno_cap']}*(1+{A['conting']})", BLACK, CUR)
r += 1
w(wsq, f"A{r}", "Value created by the interior programme", BOLD, border=TOP)
w(wsq, f"C{r}", f"=C{r-2}+C{r-1}", BOLD, CUR, border=TOP)
r += 1
w(wsq, f"A{r}", "Less: non-unit capital + its share of contingency (no rent lift)", BLACK)
w(wsq, f"C{r}", f"=-'Capex Budget'!B{CAP_TOT_ROW}*(1+{A['conting']})", BLACK, CUR)
r += 1
w(wsq, f"A{r}", "Net value created by the capital plan", BOLD, border=TOPBOT, fill=FILL_KEY)
w(wsq, f"C{r}", f"=C{r-2}+C{r-1}", BOLD, CUR, border=TOPBOT, fill=FILL_KEY)
w(wsq, f"J{r}",
  "This is the heart of the recommendation. The interior renovation is a good project on its own — a 12.7% "
  "return on cost against a 6.75% exit cap. But roughly a third of the capital budget is deferred maintenance "
  "that generates no rent, and it consumes almost all of the value the renovation creates. The operating upside "
  "the OM advertises (burning off loss to lease, occupancy to 95%, fully billing the technology and valet trash "
  "charges) is already embedded in the forecast NOI the asking price is struck against — so we would be paying "
  "today for value we have to create ourselves.", NOTE, align="wrap")
wsq.row_dimensions[r].height = 80
r += 2
wsq.merge_cells(f"A{r}:J{r}")
w(wsq, f"A{r}",
  "Cash flow convention: all capital is funded at closing, so the Year-0 outflow carries the full renovation "
  "and property capital budget and annual cash flow deducts only interest and replacement reserves. Interest is "
  "constant because the loan is interest-only and fully drawn at close.", NOTE, align="wrap")
wsq.row_dimensions[r].height = 34

# patch Annual Model references to Returns rows
for ph, row in [("{tcost}", tc), ("{debt}", QR["loan"]), ("{eq}", QR["eq"])]:
    for rr in range(1, wsy.max_row + 1):
        for cc in range(1, 9):
            cell = wsy.cell(row=rr, column=cc)
            if isinstance(cell.value, str) and ph in cell.value:
                cell.value = cell.value.replace(ph, str(row))

# ============================================================================
# 11. SENSITIVITIES
# ============================================================================
wss = wb.create_sheet("Sensitivities")
wss.sheet_view.showGridLines = False
wss.column_dimensions["A"].width = 30
for c in "BCDEFGHIJKLM":
    wss.column_dimensions[c].width = 13
title_block(wss, "Live Sensitivities and Breakeven Ladders — every cell recalculates from the Assumptions tab", "M")

AMc = "'Annual Model'"
PRICES = [10500000, 11000000, 11500000, 12000000, 12500000, 13625000, 14200000]
CAPS = [0.0600, 0.0650, 0.0675, 0.0700, 0.0750]
OTHER_CAP = f"({A['reno_total']}+'Capex Budget'!$B${CAP_TOT_ROW})*(1+{A['conting']})"
LOAN_TESTS = (f"{AMc}!C{MR['noi']}/{A['mindy']}", f"{AMc}!C{MR['noi']}/({A['mindscr']}*{A['rate']})")


def leak(yc):
    """Share of a marginal GPR dollar that reaches NOI, before the management fee."""
    return (f"(1-Assumptions!{yc}${AROW['ll']}-Assumptions!{yc}${AROW['vac']}"
            f"-Assumptions!{yc}${AROW['conc']}-Assumptions!{yc}${AROW['bd']}"
            f"-'Reno Schedule'!{yc}${NREF['dvac']})")


def cf_block(row, price_ref, dnoi_refs, cap_ref, noi1_ref, noi6_ref):
    """Write loan / equity / 6 cash-flow cells / IRR / EM for one scenario row."""
    # noi1_ref can be a compound expression, so it MUST be parenthesised before dividing
    w(wss, f"C{row}", f"=MIN(({price_ref}+{OTHER_CAP})*{A['ltc']},({noi1_ref})/{A['mindy']},"
                      f"({noi1_ref})/({A['mindscr']}*{A['rate']}))", BLACK, CUR)
    w(wss, f"D{row}", f"={price_ref}+{OTHER_CAP}+{price_ref}*{A['acqcost']}+C{row}*{A['fee']}"
                      f"+{A['fincost']}-C{row}", BLACK, CUR)
    w(wss, f"E{row}", f"=-D{row}", BLACK, CUR)
    for i in range(5):
        col = "FGHIJ"[i]
        yc = YR_COLS[i]
        f_ = f"={AMc}!{yc}{MR['cfbd']}{dnoi_refs[i]}-C{row}*{A['rate']}"
        if i == 4:
            f_ += f"+({noi6_ref})/{cap_ref}*(1-{A['cos']})-C{row}"
        w(wss, f"{col}{row}", f_, BLACK, CUR)
    w(wss, f"K{row}", f"=IRR(E{row}:J{row})", BOLD, PCT)
    w(wss, f"L{row}", f"=SUM(F{row}:J{row})/-E{row}", BLACK, MULT)


def helper_header(row, first, second):
    for col, h in [("A", first), ("B", second), ("C", "Loan"), ("D", "Equity"), ("E", "Year 0"),
                   ("F", "Year 1"), ("G", "Year 2"), ("H", "Year 3"), ("I", "Year 4"), ("J", "Year 5"),
                   ("K", "Levered IRR"), ("L", "Equity Mult.")]:
        w(wss, f"{col}{row}", h, BOLD, fill=FILL_SUB, align="wrap")
    return row + 1


# ---------------------------------------------------------------- Grid 1 ---
r = 5
r = section(wss, r, "HELPER BLOCK 1 — PURCHASE PRICE x EXIT CAP", "M")
r = helper_header(r, "Purchase Price", "Exit Cap")
h1 = r
for p_ in PRICES:
    for cap in CAPS:
        w(wss, f"A{r}", p_, BLUE, CUR)
        w(wss, f"B{r}", cap, BLUE, PCT2, align="center")
        cf_block(r, f"$A{r}", [""] * 5, f"$B{r}", f"{AMc}!C{MR['noi']}", f"{AMc}!{EXIT_COL}{MR['noi']}")
        r += 1

r += 1
r = section(wss, r, "GRID 1A — LEVERED IRR: PURCHASE PRICE x EXIT CAP", "H")
w(wss, f"A{r}", "Purchase Price \\ Exit Cap", BOLD, fill=FILL_SUB, align="wrap")
for j, cap in enumerate(CAPS):
    w(wss, f"{'BCDEF'[j]}{r}", cap, BOLD, PCT2, fill=FILL_SUB, align="center")
w(wss, f"G{r}", "Price / Unit", BOLD, fill=FILL_SUB, align="center")
w(wss, f"H{r}", "Going-in Cap", BOLD, fill=FILL_SUB, align="center")
r += 1
for i, p_ in enumerate(PRICES):
    w(wss, f"A{r}", p_, BLACK, CUR)
    for j, cap in enumerate(CAPS):
        src = h1 + i * len(CAPS) + j
        hl = abs(cap - 0.0675) < 1e-9 and p_ in (13625000,)
        w(wss, f"{'BCDEF'[j]}{r}", f"=K{src}", BOLD if p_ == 13625000 else BLACK, PCT,
          align="center", fill=FILL_KEY if hl else None, border=BOX)
    w(wss, f"G{r}", f"=A{r}/{A['units']}", ITAL, CUR)
    w(wss, f"H{r}", f"={AMc}!C{MR['noi']}/A{r}", ITAL, PCT2)
    r += 1
r += 1
wss.merge_cells(f"A{r}:H{r}")
w(wss, f"A{r}",
  "The highlighted row is the price the asset actually traded at on 12 June 2026 ($13,625,000 / $115,466 per "
  "unit, Marion County deed, assessor validity flag 'Y'). Our bid is the highest price in this ladder that "
  "still clears a 14% levered IRR and a 1.80x multiple.", NOTE, align="wrap")
wss.row_dimensions[r].height = 34
r += 2

r = section(wss, r, "GRID 1B — EQUITY MULTIPLE: PURCHASE PRICE x EXIT CAP", "F")
w(wss, f"A{r}", "Purchase Price \\ Exit Cap", BOLD, fill=FILL_SUB, align="wrap")
for j, cap in enumerate(CAPS):
    w(wss, f"{'BCDEF'[j]}{r}", cap, BOLD, PCT2, fill=FILL_SUB, align="center")
r += 1
for i, p_ in enumerate(PRICES):
    w(wss, f"A{r}", p_, BLACK, CUR)
    for j, cap in enumerate(CAPS):
        src = h1 + i * len(CAPS) + j
        w(wss, f"{'BCDEF'[j]}{r}", f"=L{src}", BLACK, MULT, align="center", border=BOX)
    r += 1
r += 2

# ---------------------------------------------------------------- Grid 2 ---
GROWTHS = [0.010, 0.015, 0.020, 0.025, 0.030, 0.035]
r = section(wss, r, "HELPER BLOCK 2 — FLAT MARKET RENT GROWTH x EXIT CAP", "M")
w(wss, f"A{r}", "Applies a single flat growth rate in place of the 1.5 / 2.0 / 2.5 / 3.0 / 3.0 base path. "
                "Gross potential rent in year y scales by (1+g)^(y-0.5) divided by the base mid-year factor; "
                "the change flows to NOI net of loss to lease, vacancy, downtime, concessions, bad debt and "
                "the management fee.", NOTE, align="wrap")
wss.merge_cells(f"A{r}:M{r}")
wss.row_dimensions[r].height = 34
r += 1
r = helper_header(r, "Flat Rent Growth", "Exit Cap")
h2 = r
for g_ in GROWTHS:
    for cap in CAPS:
        w(wss, f"A{r}", g_, BLUE, PCT, align="center")
        w(wss, f"B{r}", cap, BLUE, PCT2, align="center")
        dnoi, noi_y = [], {}
        for i in range(5):
            yc = YR_COLS[i]
            scale = f"((1+$A{r})^({i+1}-0.5)/{AMc}!{yc}{MR['fac']}-1)"
            dg = f"({AMc}!{yc}{MR['gpr']}*{scale})"
            dnoi.append(f"+{dg}*{leak(yc)}*(1-{A['e_mgmt']})")
            noi_y[i] = f"{AMc}!{yc}{MR['noi']}+{dg}*{leak(yc)}*(1-{A['e_mgmt']})"
        scale6 = f"((1+$A{r})^(6-0.5)/{AMc}!{EXIT_COL}{MR['fac']}-1)"
        leak6 = (f"(1-Assumptions!G${AROW['ll']}-Assumptions!G${AROW['vac']}"
                 f"-Assumptions!G${AROW['conc']}-Assumptions!G${AROW['bd']})")
        noi6 = (f"{AMc}!{EXIT_COL}{MR['noi']}+{AMc}!{EXIT_COL}{MR['gpr']}*{scale6}"
                f"*{leak6}*(1-{A['e_mgmt']})")
        cf_block(r, f"{A['price']}", dnoi, f"$B{r}", noi_y[0], noi6)
        r += 1

r += 1
r = section(wss, r, "GRID 2 — LEVERED IRR: FLAT RENT GROWTH x EXIT CAP (at our bid)", "G")
w(wss, f"A{r}", "Flat Rent Growth \\ Exit Cap", BOLD, fill=FILL_SUB, align="wrap")
for j, cap in enumerate(CAPS):
    w(wss, f"{'BCDEF'[j]}{r}", cap, BOLD, PCT2, fill=FILL_SUB, align="center")
w(wss, f"G{r}", "Market evidence", BOLD, fill=FILL_SUB, align="center")
r += 1
NOTES_G = {0.010: "1.1% actual YoY, 1Q-26", 0.015: "our Year 1", 0.020: "our Year 2",
           0.025: "our Year 3", 0.030: "our Years 4-5", 0.035: "above trend"}
for i, g_ in enumerate(GROWTHS):
    w(wss, f"A{r}", g_, BLACK, PCT)
    for j, cap in enumerate(CAPS):
        src = h2 + i * len(CAPS) + j
        w(wss, f"{'BCDEF'[j]}{r}", f"=K{src}", BLACK, PCT, align="center", border=BOX)
    w(wss, f"G{r}", NOTES_G[g_], ITAL, align="center")
    r += 1
r += 2

# ------------------------------------------------------- Breakeven ladders ---
r = section(wss, r, "BREAKEVEN LADDERS — WHERE EACH DRIVER CROSSES THE HURDLE", "M")
w(wss, f"A{r}", "Each ladder flexes ONE driver at our bid price and the base exit cap, holding everything "
                "else at the base case. Read down each block to find where the levered IRR crosses 10%, 12% "
                "and 14%. Every cell is live.", NOTE, align="wrap")
wss.merge_cells(f"A{r}:M{r}")
wss.row_dimensions[r].height = 28
r += 2

LADDERS = []

# (a) blended renovation premium
r = section(wss, r, "LADDER A — BLENDED RENOVATION PREMIUM ($/unit/month)", "M")
r = helper_header(r, "Blended Premium", "(base)")
la = r
PREMS = [100, 125, 140, 155, 170, 190]
for pm in PREMS:
    w(wss, f"A{r}", pm, BLUE, CUR)
    w(wss, f"B{r}", f"={A['prem_blend']}", GREEN, CUR)
    ratio = f"($A{r}/$B{r})"
    dnoi, noi_y = [], {}
    for i in range(5):
        yc = YR_COLS[i]
        dg = f"({AMc}!{yc}{MR['gpr_r']}*({ratio}-1))"
        dnoi.append(f"+{dg}*{leak(yc)}*(1-{A['e_mgmt']})")
        noi_y[i] = f"{AMc}!{yc}{MR['noi']}+{dg}*{leak(yc)}*(1-{A['e_mgmt']})"
    leak6 = (f"(1-Assumptions!G${AROW['ll']}-Assumptions!G${AROW['vac']}"
             f"-Assumptions!G${AROW['conc']}-Assumptions!G${AROW['bd']})")
    noi6 = (f"{AMc}!{EXIT_COL}{MR['noi']}+{AMc}!{EXIT_COL}{MR['gpr_r']}*({ratio}-1)"
            f"*{leak6}*(1-{A['e_mgmt']})")
    cf_block(r, f"{A['price']}", dnoi, f"{A['exitcap']}", noi_y[0], noi6)
    w(wss, f"M{r}", f"=$A{r}*12/{A['cost_blend']}", ITAL, PCT)
    r += 1
w(wss, f"M{la-1}", "Return on Cost", BOLD, fill=FILL_SUB, align="wrap")
r += 1

# (b) blended renovation cost
r = section(wss, r, "LADDER B — BLENDED RENOVATION COST ($/unit)", "M")
r = helper_header(r, "Blended Cost", "(base)")
lb = r
COSTS = [11000, 12500, 14407, 16000, 18000, 20000]
for cst in COSTS:
    w(wss, f"A{r}", cst, BLUE, CUR)
    w(wss, f"B{r}", f"={A['cost_blend']}", GREEN, CUR)
    other = f"($A{r}*{A['units']}+'Capex Budget'!$B${CAP_TOT_ROW})*(1+{A['conting']})"
    w(wss, f"C{r}", f"=MIN(({A['price']}+{other})*{A['ltc']},{LOAN_TESTS[0]},{LOAN_TESTS[1]})", BLACK, CUR)
    w(wss, f"D{r}", f"={A['price']}+{other}+{A['price']}*{A['acqcost']}+C{r}*{A['fee']}"
                    f"+{A['fincost']}-C{r}", BLACK, CUR)
    w(wss, f"E{r}", f"=-D{r}", BLACK, CUR)
    for i in range(5):
        col = "FGHIJ"[i]
        f_ = f"={AMc}!{YR_COLS[i]}{MR['cfbd']}-C{r}*{A['rate']}"
        if i == 4:
            f_ += f"+{AMc}!{EXIT_COL}{MR['noi']}/{A['exitcap']}*(1-{A['cos']})-C{r}"
        w(wss, f"{col}{r}", f_, BLACK, CUR)
    w(wss, f"K{r}", f"=IRR(E{r}:J{r})", BOLD, PCT)
    w(wss, f"L{r}", f"=SUM(F{r}:J{r})/-E{r}", BLACK, MULT)
    w(wss, f"M{r}", f"={A['prem_blend']}*12/$A{r}", ITAL, PCT)
    r += 1
w(wss, f"M{lb-1}", "Return on Cost", BOLD, fill=FILL_SUB, align="wrap")
r += 1

# (c) real estate taxes — the reassessment question
r = section(wss, r, "LADDER C — REAL ESTATE TAX REASSESSMENT (Year-5 bill)", "M")
r = helper_header(r, "Year-5 Tax Bill", "Implied AV")
lc = r
TAXES = [(137000, "no growth from the 2025 actual"),
         (160300, "base case — actual bill grown 4%"),
         (185000, "AV to ~55% of the trade price"),
         (238700, "AV to 69% of the trade price (OM comp average)"),
         (294000, "AV to 85% of the trade price")]
for tx, note in TAXES:
    w(wss, f"A{r}", tx, BLUE, CUR)
    w(wss, f"B{r}", f"=A{r}/0.024861/{A['actual']}", BLACK, PCT)
    dnoi, noi_y = [], {}
    for i in range(5):
        yc = YR_COLS[i]
        step = f"(($A{r}-Assumptions!G${AROW['e_tax']})*{i+1}/5)"
        dnoi.append(f"-{step}")
        noi_y[i] = f"{AMc}!{yc}{MR['noi']}-{step}"
    noi6 = f"{AMc}!{EXIT_COL}{MR['noi']}-($A{r}-Assumptions!G${AROW['e_tax']})*1.04"
    cf_block(r, f"{A['price']}", dnoi, f"{A['exitcap']}", noi_y[0], noi6)
    w(wss, f"M{r}", note, ITAL, align="wrap")
    r += 1
w(wss, f"M{lc-1}", "Scenario", BOLD, fill=FILL_SUB, align="wrap")
r += 1
wss.merge_cells(f"A{r}:M{r}")
w(wss, f"A{r}",
  "Marion County parcel 8000002 has carried a gross assessed value between $5.1mm and $6.6mm every year since "
  "2012, and the last assessment change was recorded 03/15/2022. The 2006 sale at $4.4mm did not reset it. "
  "Indiana values apartments at the LOWEST of the income, market and cost approaches, which is why assessed "
  "value has never tracked trade prices here. That history is the reason the base case grows the actual "
  "$129,648 bill at 4% rather than stepping it to a share of the purchase price — but the risk is real and "
  "this ladder prices it.", NOTE, align="wrap")
wss.row_dimensions[r].height = 56

# ============================================================================
# 11b. TAX RECORD
# ============================================================================
wtx = wb.create_sheet("Tax Record")
wtx.sheet_view.showGridLines = False
for col, wd in [("A", 12), ("B", 15), ("C", 15), ("D", 15), ("E", 12), ("F", 15), ("G", 14), ("H", 60)]:
    wtx.column_dimensions[col].width = wd
title_block(wtx, "Marion County Parcel 8000002 — 25-Year Assessment and Tax Record", "H")

r = 5
r = section(wtx, r, "PARCEL", "H")
for lbl, val in [
    ("Parcel Number", "8000002"), ("State Parcel Number", "49-07-04-115-001.000-800"),
    ("Owner of Record", "4020 MONACO DRIVE OWNER LLC"), ("Use Code", "403 (apartments)"),
    ("Tax District", "800"), ("Year Built", "1974"), ("Acreage", "9.40"),
    ("Deed Type", "Special Warranty Deed"), ("Deed Date", "12 June 2026"),
    ("Recorded", "26 June 2026"), ("Last Assessment Change", "15 March 2022"),
]:
    w(wtx, f"A{r}", lbl, BLACK)
    w(wtx, f"B{r}", val, BLUE)
    r += 1

r += 1
r = section(wtx, r, "ASSESSMENT AND TAX HISTORY", "H")
for col, h in [("A", "Tax Year"), ("B", "Land AV"), ("C", "Improvements"), ("D", "Gross AV"),
               ("E", "Tax Rate"), ("F", "Net Annual Tax"), ("G", "Tax / Unit")]:
    w(wtx, f"{col}{r}", h, BOLD, fill=FILL_SUB, align="center")
w(wtx, f"H{r}", "Note", BOLD, fill=FILL_SUB)
r += 1
HIST = [
    (2025, 1085500, 4462300, 5547800, .024861, 129648.42, "Gross AV less a $332,868 '2% deduction' = $5,214,932 net"),
    (2024, 1085500, 4021100, 5106600, .0254, 129656.56, ""),
    (2023, 1085500, 5471600, 6557100, .0254, 166425.76, "25-year peak assessed value"),
    (2022, 1085500, 4678300, 5763800, .0257, 147893.34, "Last assessment change recorded 15 March 2022"),
    (2021, 986800, 4337100, 5323900, .0270, 139656.54, ""),
    (2020, 986800, 4343900, 5330700, .0259, 132409.24, ""),
    (2019, 986800, 4654600, 5641400, .0244, 131783.10, ""),
    (2018, 986800, 4654600, 5641400, .0239, 128832.64, ""),
    (2017, 986800, 4453700, 5440500, .0237, 122802.98, ""),
    (2016, 986800, 4210700, 5197500, .0225, 109667.24, ""),
    (2015, 986800, 4216800, 5203600, .0207, 107771.76, ""),
    (2014, 986800, 4310000, 5296800, .0206, 109172.34, ""),
    (2013, 986900, 4310000, 5296900, .0220, 110175.50, ""),
    (2012, 986900, 4310000, 5296900, .0202, 106806.68, "Assessed value has been range-bound ever since"),
    (2011, 511800, 2988300, 3500100, .0195, 68118.96, ""),
    (2010, 511800, 2988300, 3500100, .0191, 67005.92, ""),
    (2009, 511800, 3536200, 4048000, .0175, 70937.14, ""),
    (2008, 511800, 3536300, 4048100, .0181, 73396.10, ""),
    (2007, 511800, 3536200, 4048000, .0239, 74739.82, "AV FELL the year after the March 2006 sale"),
    (2006, 694000, 4562800, 5256800, .0245, 97947.12, "Property sold 28 March 2006 for $4,400,000 (flagged 'N')"),
    (2005, 257600, 3963400, 4221000, .0256, 78113.26, ""),
    (2004, 257600, 3963400, 4221000, .0248, 77326.46, ""),
    (2003, 257600, 3963400, 4221000, .0228, 67727.90, ""),
    (2002, 257600, 3963400, 4221000, .0247, 75149.16, ""),
    (2001, 329000, 2313000, 2642000, .0326, 73755.90, ""),
    (2000, 141000, 771000, 912000, .0936, 73385.76, "Pre-2002 assessment methodology; not comparable"),
]
h0 = r
for yr, land, imp, gross, rate, tax, note in HIST:
    w(wtx, f"A{r}", str(yr), BLACK, align="center")
    w(wtx, f"B{r}", land, BLUE, CUR)
    w(wtx, f"C{r}", imp, BLUE, CUR)
    w(wtx, f"D{r}", f"=B{r}+C{r}", BLACK, CUR)
    w(wtx, f"E{r}", rate, BLUE, PCT2, align="center")
    w(wtx, f"F{r}", tax, BLUE, CUR)
    w(wtx, f"G{r}", f"=F{r}/{A['units']}", BLACK, CUR)
    if note:
        w(wtx, f"H{r}", note, NOTE, align="wrap")
    r += 1
hN = r - 1

r += 1
r = section(wtx, r, "WHAT THE RECORD IMPLIES FOR THE UNDERWRITING", "H")
stats = [
    ("Gross AV, 2012", f"=D{h0+13}", CUR, ""),
    ("Gross AV, 2025", f"=D{h0}", CUR, ""),
    ("  Compound annual growth, 2012-2025", f"=(D{h0}/D{h0+13})^(1/13)-1", PCT2,
     "Assessed value has effectively stood still for thirteen years."),
    ("Net tax, 2012", f"=F{h0+13}", CUR, ""),
    ("Net tax, 2025", f"=F{h0}", CUR, ""),
    ("  Compound annual growth, 2012-2025", f"=(F{h0}/F{h0+13})^(1/13)-1", PCT2,
     "The bill has risen with the RATE, not the assessment."),
    ("Highest gross AV on record (2023)", f"=MAX(D{h0}:D{hN})", CUR, ""),
    ("2025 gross AV as a % of the 2026 trade price", f"=D{h0}/{A['actual']}", PCT,
     "The OM's own comp study assumes assessed value runs at about 69% of sale price after a trade. "
     "This parcel is at 41% BEFORE any reassessment."),
    ("Year-1 tax underwritten", f"=Assumptions!C${AROW['e_tax']}", CUR, ""),
    ("Year-5 tax underwritten", f"=Assumptions!G${AROW['e_tax']}", CUR,
     "Base case grows the actual bill at 4% a year — well above the 1.5% the record shows."),
    ("Year-5 tax if AV reset to 69% of the trade price", f"={A['actual']}*0.69*0.024861", CUR,
     "Priced as Ladder C on the Sensitivities tab, not in the base case."),
]
for lbl, f, fmt, note in stats:
    w(wtx, f"A{r}", lbl, BLACK)
    wtx.merge_cells(f"A{r}:C{r}")
    w(wtx, f"D{r}", f, BOLD if not lbl.startswith("  ") else BLACK, fmt)
    if note:
        w(wtx, f"H{r}", note, NOTE, align="wrap")
        wtx.row_dimensions[r].height = 28
    r += 1

r += 1
wtx.merge_cells(f"A{r}:H{r}")
w(wtx, f"A{r}",
  "Why this matters. Indiana values apartments at the LOWEST of the income, market and cost approaches, and "
  "this record shows the result: the 2006 sale at $4,400,000 was followed by assessed value FALLING the next "
  "year, and assessed value has stayed between $5.1mm and $6.6mm every year since 2012 while the property "
  "traded twice. An earlier draft of this model stepped taxes from $140,000 to $210,000 on the assumption "
  "that a sale resets the assessment. The public record does not support that, and correcting it is worth "
  "roughly 1.5 points of levered IRR. The risk has not been dismissed — it is priced explicitly in Ladder C.",
  NOTE, align="wrap")
wtx.row_dimensions[r].height = 60

# ============================================================================
# 12. DD QUESTIONS
# ============================================================================
wsd = wb.create_sheet("DD Questions")
wsd.sheet_view.showGridLines = False
wsd.column_dimensions["A"].width = 6
wsd.column_dimensions["B"].width = 20
wsd.column_dimensions["C"].width = 74
wsd.column_dimensions["D"].width = 60
wsd.column_dimensions["E"].width = 12
title_block(wsd, "Due Diligence & Seller / Broker Question List — with the answer assumed in this model", "E")

import json as _json
DD = [(sec, [tuple(q) for q in items])
      for sec, items in _json.load(open(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                                     "dd_questions.json")))]

r = 5
w(wsd, f"A{r}", "#", BOLD, fill=FILL_SUB, align="center")
w(wsd, f"B{r}", "Category", BOLD, fill=FILL_SUB)
w(wsd, f"C{r}", "Question for the seller / broker", BOLD, fill=FILL_SUB)
w(wsd, f"D{r}", "Reasonable answer assumed in this model", BOLD, fill=FILL_SUB)
w(wsd, f"E{r}", "Priority", BOLD, fill=FILL_SUB, align="center")
r += 1
n = 0
for cat, items in DD:
    w(wsd, f"A{r}", "", BLACK)
    wsd.merge_cells(f"B{r}:E{r}")
    w(wsd, f"B{r}", cat.upper(), BOLD, fill=FILL_SUB)
    for col in ["C", "D", "E"]:
        wsd[f"{col}{r}"].fill = FILL_SUB
    r += 1
    for q, ans, pri in items:
        n += 1
        w(wsd, f"A{r}", n, BLACK, NUM, align="center", border=BOX)
        w(wsd, f"B{r}", cat.split(". ", 1)[-1], NOTE, align="wrap", border=BOX)
        w(wsd, f"C{r}", q, BLACK, align="wrap", border=BOX)
        w(wsd, f"D{r}", ans, NOTE, align="wrap", border=BOX)
        w(wsd, f"E{r}", pri, BOLD if pri == "High" else BLACK, align="center", border=BOX,
          fill=FILL_KEY if pri == "High" else None)
        wsd.row_dimensions[r].height = max(30, 13 * (max(len(q), len(ans)) // 70 + 1))
        r += 1
w(wsd, f"A{r+1}", f"{n} questions. 'High' priority items are those where a different answer changes the "
                  "underwriting or the recommendation.", NOTE)

# ----------------------------------------------------------------------------
# Print set-up and freeze panes — so the workbook is usable on screen and on paper
# ----------------------------------------------------------------------------
FREEZE = {
    "Assumptions": "B5", "Rent Roll": "B6", "Unit Mix": "B6", "T-12 Recast": "B6",
    "Other Income": "B6", "Capex Budget": "B6", "Reno Schedule": "B6",
    "Annual Model": "C7", "Returns": "B6", "Sensitivities": "B6", "Tax Record": "B6", "DD Questions": "C6",
}
for sheet in wb.worksheets:
    ps = sheet.page_setup
    ps.orientation = "landscape"
    sheet.page_margins.left = sheet.page_margins.right = 0.4
    sheet.page_margins.top = sheet.page_margins.bottom = 0.5
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    ps.fitToWidth = 1
    ps.fitToHeight = 0
    sheet.print_options.horizontalCentered = False
    if sheet.title in FREEZE:
        sheet.freeze_panes = FREEZE[sheet.title]
    if sheet.title in ("Rent Roll", "DD Questions", "Reno Schedule"):
        sheet.print_title_rows = "5:5" if sheet.title != "Reno Schedule" else "5:5"

wb.save(OUT)
print("wrote", OUT)
print("questions:", n)
